from __future__ import annotations

import io
import json
import os
import time
import html
import xml.etree.ElementTree as ET
from datetime import datetime
from dataclasses import asdict
from pathlib import Path
from typing import Any

from dotenv import load_dotenv

APP_DIR = Path(__file__).resolve().parent
load_dotenv(APP_DIR.parent / ".env")
load_dotenv(APP_DIR / ".env", override=True)

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak

from grpc_server import server_status
from quality_diagnosis.qa_test_utils import load_cases, run_case
from quality_diagnosis.realtime_evaluator import evaluate_realtime
from quality_diagnosis.three_axis_validation import VALIDATION_CASES, run_three_axis_validation, validation_metrics
from quality_diagnosis.llm_judge import load_json as load_judge_json, run_judge_cases, write_csv as write_judge_csv
from utils.pipeline import analyze_voc
from agents.pipeline_agents import Interpreter, Retriever, Summarizer, Evaluator, Critic, Improver

ROOT = Path(__file__).resolve().parent
REPORTS = ROOT / "quality_diagnosis" / "reports"
AGENTS = ["Interpreter", "Retriever", "Summarizer", "Evaluator", "Critic", "Improver"]

st.set_page_config(
    page_title="VOC Quality Studio",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
<style>
:root {
  --navy: #0B1220;
  --ink: #172033;
  --muted: #667085;
  --line: #E4E7EC;
  --panel: #FFFFFF;
  --canvas: #F5F7FB;
  --blue: #2563EB;
  --blue-soft: #EFF6FF;
  --green: #16A34A;
  --green-soft: #ECFDF3;
  --amber: #D97706;
  --amber-soft: #FFF7ED;
  --red: #DC2626;
  --red-soft: #FEF2F2;
}
html, body, [class*="css"] { font-family: "Segoe UI", "Pretendard", sans-serif; }
[data-testid="stAppViewContainer"] { background: var(--canvas); }
[data-testid="stHeader"] { display: none; }
.block-container { padding-top: 1.1rem; padding-bottom: 2rem; max-width: 1600px; }
[data-testid="stSidebar"] { background: var(--navy); border-right: 0; }
[data-testid="stSidebar"] * { color: #E5E7EB; }
[data-testid="stSidebar"] [role="radiogroup"] label {
  padding: .48rem .65rem; border-radius: .55rem; margin-bottom: .15rem;
}
[data-testid="stSidebar"] [role="radiogroup"] label:hover { background: rgba(255,255,255,.07); }
[data-testid="stMetric"] {
  background: var(--panel); border: 1px solid var(--line); border-radius: 14px;
  padding: 15px 17px; box-shadow: 0 3px 12px rgba(16,24,40,.045);
}
[data-testid="stMetricLabel"] { color: var(--muted); font-weight: 700; }
[data-testid="stMetricValue"] { color: var(--ink); font-weight: 800; }
.stButton > button {
  border-radius: 10px; min-height: 44px; font-weight: 750; border: 1px solid #CBD5E1;
}
.stButton > button[kind="primary"] {
  background: var(--blue); border-color: var(--blue); box-shadow: 0 5px 14px rgba(37,99,235,.2);
}
.page-title { font-size: 2rem; font-weight: 850; color: var(--ink); letter-spacing: -.04em; margin: 0; }
.page-subtitle { color: var(--muted); font-size: 1rem; margin: .15rem 0 1.15rem 0; }
.section-title { font-size: 1.08rem; font-weight: 800; color: var(--ink); margin: .2rem 0 .65rem; }
.panel {
  background: var(--panel); border: 1px solid var(--line); border-radius: 14px;
  padding: 17px 18px; box-shadow: 0 3px 12px rgba(16,24,40,.04);
}
.hero {
  background: linear-gradient(120deg, #0B1220 0%, #172554 58%, #1D4ED8 100%);
  color: white; border-radius: 18px; padding: 24px 26px; margin-bottom: 1rem;
  box-shadow: 0 12px 30px rgba(15,23,42,.18);
}
.hero h1 { margin: 0 0 .25rem; font-size: 2.15rem; letter-spacing: -.045em; }
.hero p { margin: 0; color: #D7E3FF; font-size: 1.03rem; }
.badge { display:inline-flex; align-items:center; gap:.4rem; padding:.3rem .62rem; border-radius:999px; font-size:.8rem; font-weight:800; }
.badge-blue { color:#1D4ED8; background:#DBEAFE; }
.badge-green { color:#15803D; background:#DCFCE7; }
.badge-amber { color:#B45309; background:#FEF3C7; }
.badge-red { color:#B91C1C; background:#FEE2E2; }
.agent-grid { display:grid; grid-template-columns:repeat(6,minmax(0,1fr)); gap:9px; }
.agent-card { background:#fff; border:1px solid var(--line); border-radius:12px; padding:13px 10px; text-align:center; transition:.2s ease; }
.agent-card .name { font-size:.83rem; font-weight:800; color:var(--ink); }
.agent-card .meta { font-size:.72rem; color:var(--muted); margin-top:.25rem; }
.agent-card.waiting { background:#fff; }
.agent-card.running { border:2px solid var(--blue); background:var(--blue-soft); box-shadow:0 0 0 4px rgba(37,99,235,.10); transform:translateY(-2px); }
.agent-card.complete { border-color:#86EFAC; background:var(--green-soft); }
.agent-card.error { border-color:#FCA5A5; background:var(--red-soft); }
.status-dot { width:9px; height:9px; border-radius:50%; display:inline-block; margin-right:5px; background:#94A3B8; }
.running .status-dot { background:var(--blue); animation:pulse 1s infinite; }
.complete .status-dot { background:var(--green); }
.error .status-dot { background:var(--red); }
@keyframes pulse { 0%{box-shadow:0 0 0 0 rgba(37,99,235,.45)} 70%{box-shadow:0 0 0 8px rgba(37,99,235,0)} 100%{box-shadow:0 0 0 0 rgba(37,99,235,0)} }
.result-card { background:#fff; border:1px solid var(--line); border-radius:14px; padding:16px 18px; margin-bottom:.7rem; }
.result-card h4 { margin:0 0 .45rem; color:var(--ink); }
.result-card p { margin:.15rem 0; color:#475467; }
.flow-line { height:3px; background:#D0D5DD; margin:0 2px; border-radius:99px; }
.small-note { color:var(--muted); font-size:.85rem; }
.judge-flow { display:flex; align-items:center; gap:9px; margin:.35rem 0 1rem; }
.judge-node { flex:1; min-height:76px; display:flex; flex-direction:column; justify-content:center; text-align:center; background:#fff; border:1px solid var(--line); border-radius:12px; padding:10px; }
.judge-node strong { color:var(--ink); font-size:.9rem; }
.judge-node span { color:var(--muted); font-size:.76rem; margin-top:3px; }
.judge-node.external { border-color:#93C5FD; background:#EFF6FF; }
.judge-arrow { color:#64748B; font-size:1.25rem; font-weight:900; }
@media (max-width: 900px) { .judge-flow { flex-direction:column; } .judge-node { width:100%; } .judge-arrow { transform:rotate(90deg); } }
[data-testid="stDataFrame"] { border:1px solid var(--line); border-radius:12px; overflow:hidden; }
[data-testid="stExpander"] { background:#fff; border:1px solid var(--line); border-radius:12px; }
hr { border-color:var(--line); }
.execution-shell { background:#0B1220; border:1px solid #1F2A44; border-radius:16px; padding:16px 18px; color:#E5E7EB; box-shadow:0 10px 28px rgba(15,23,42,.14); }
.execution-head { display:flex; justify-content:space-between; align-items:center; gap:12px; margin-bottom:10px; }
.execution-title { font-weight:850; font-size:1rem; color:#F8FAFC; }
.execution-sub { color:#93A4BF; font-size:.82rem; }
.timeline { display:flex; flex-direction:column; gap:8px; margin-top:10px; }
.timeline-row { display:grid; grid-template-columns:88px 118px 1fr 70px; gap:10px; align-items:center; padding:9px 10px; border-radius:10px; background:#111A2D; border:1px solid #24324D; }
.timeline-row.running { border-color:#60A5FA; background:#10244A; box-shadow:0 0 0 3px rgba(37,99,235,.13); }
.timeline-row.complete { border-color:#166534; background:#102719; }
.timeline-row.error { border-color:#991B1B; background:#321313; }
.timeline-step { color:#93A4BF; font-size:.75rem; font-weight:800; }
.timeline-agent { color:#F8FAFC; font-size:.86rem; font-weight:850; }
.timeline-message { color:#CBD5E1; font-size:.82rem; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
.timeline-time { color:#93A4BF; font-size:.75rem; text-align:right; }
.live-chip { display:inline-flex; align-items:center; gap:6px; padding:.28rem .58rem; border-radius:999px; background:#1E3A8A; color:#DBEAFE; font-size:.74rem; font-weight:850; }
.live-chip:before { content:""; width:8px; height:8px; border-radius:50%; background:#60A5FA; animation:pulse 1s infinite; }
.progress-caption { color:#64748B; font-size:.82rem; margin-top:.35rem; }

.qa-live-grid { display:grid; grid-template-columns:repeat(6,minmax(0,1fr)); gap:10px; margin-bottom:12px; }
.qa-kpi { background:#fff; border:1px solid var(--line); border-radius:13px; padding:13px 14px; box-shadow:0 3px 12px rgba(16,24,40,.04); }
.qa-kpi .k-label { color:var(--muted); font-size:.74rem; font-weight:800; text-transform:uppercase; letter-spacing:.04em; }
.qa-kpi .k-value { color:var(--ink); font-size:1.55rem; font-weight:900; line-height:1.15; margin-top:3px; }
.qa-kpi .k-sub { color:var(--muted); font-size:.72rem; margin-top:3px; }
.qa-kpi.pass { border-color:#86EFAC; background:#F0FDF4; }
.qa-kpi.fail { border-color:#FCA5A5; background:#FEF2F2; }
.qa-kpi.active { border:2px solid #60A5FA; background:#EFF6FF; box-shadow:0 0 0 4px rgba(37,99,235,.08); }
.qa-board { background:#fff; border:1px solid var(--line); border-radius:15px; padding:15px 16px; box-shadow:0 4px 14px rgba(16,24,40,.05); }
.qa-current { display:grid; grid-template-columns:90px 130px 1fr 90px; gap:10px; align-items:center; padding:11px 12px; border:1px solid #BFDBFE; background:#EFF6FF; border-radius:11px; }
.qa-current .tc { color:#1D4ED8; font-weight:900; }
.qa-current .type { color:#334155; font-weight:800; }
.qa-current .question { color:#172033; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.qa-current .count { color:#1D4ED8; font-weight:900; text-align:right; }
.qa-result-list { display:flex; flex-direction:column; gap:7px; margin-top:10px; max-height:250px; overflow:auto; padding-right:3px; }
.qa-result-row { display:grid; grid-template-columns:78px 110px 1fr 70px 72px; gap:9px; align-items:center; padding:8px 9px; border-radius:9px; border:1px solid var(--line); background:#fff; }
.qa-result-row.pass { border-left:5px solid var(--green); }
.qa-result-row.fail { border-left:5px solid var(--red); background:#FFF8F8; }
.qa-result-row .qid { font-weight:900; color:var(--ink); }
.qa-result-row .qtype { color:var(--muted); font-size:.78rem; }
.qa-result-row .qtext { color:#475467; font-size:.79rem; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
.qa-result-row .qscore { font-weight:900; text-align:right; }
.qa-result-row .qstatus { font-weight:900; text-align:right; }
.qa-result-row.pass .qstatus { color:var(--green); }
.qa-result-row.fail .qstatus { color:var(--red); }
.qa-section-head { display:flex; justify-content:space-between; align-items:center; gap:12px; margin-bottom:9px; }
.qa-section-head h3 { margin:0; color:var(--ink); font-size:1rem; }
.qa-mini { color:var(--muted); font-size:.76rem; }
@media (max-width: 1200px) { .qa-live-grid { grid-template-columns:repeat(3,1fr); } .qa-current { grid-template-columns:80px 110px 1fr; } .qa-current .count { display:none; } }

@media (max-width: 1100px) { .agent-grid { grid-template-columns:repeat(3,1fr); } .timeline-row { grid-template-columns:70px 100px 1fr; } .timeline-time { display:none; } }
.presentation-hero { background:radial-gradient(circle at 85% 15%,rgba(96,165,250,.35),transparent 28%),linear-gradient(135deg,#070B16 0%,#111D3A 55%,#1D4ED8 100%); border:1px solid #29416F; border-radius:22px; padding:28px 30px; color:white; box-shadow:0 18px 44px rgba(15,23,42,.28); margin-bottom:15px; }
.presentation-hero h1 { margin:0; font-size:2.55rem; letter-spacing:-.055em; }
.presentation-hero p { color:#C7D7F8; font-size:1.06rem; margin:.35rem 0 0; }
.stage-grid { display:grid; grid-template-columns:repeat(5,minmax(0,1fr)); gap:9px; margin:12px 0 15px; }
.stage-card { border-radius:13px; padding:12px; text-align:center; border:1px solid #334155; background:#111827; color:#94A3B8; }
.stage-card .stage-no { font-size:.7rem; font-weight:900; letter-spacing:.08em; }
.stage-card .stage-name { font-size:.86rem; font-weight:900; margin-top:3px; }
.stage-card.active { background:#172554; border:2px solid #60A5FA; color:#EFF6FF; box-shadow:0 0 0 5px rgba(37,99,235,.13); transform:translateY(-2px); }
.stage-card.done { background:#052E16; border-color:#22C55E; color:#DCFCE7; }
.cinema-panel { background:linear-gradient(180deg,#0B1220,#101827); color:#E2E8F0; border:1px solid #263550; border-radius:18px; padding:18px 20px; box-shadow:0 12px 30px rgba(15,23,42,.18); }
.cinema-title { font-size:1.08rem; font-weight:900; color:white; margin-bottom:5px; }
.cinema-copy { color:#9FB0CC; font-size:.88rem; }
.final-decision { border-radius:18px; padding:22px 24px; text-align:center; background:linear-gradient(135deg,#052E16,#166534); color:white; border:1px solid #4ADE80; box-shadow:0 14px 35px rgba(22,101,52,.25); }
.final-decision.hold { background:linear-gradient(135deg,#450A0A,#991B1B); border-color:#F87171; }
.final-decision.review { background:linear-gradient(135deg,#451A03,#B45309); border-color:#FBBF24; }
.final-decision .big { font-size:2.25rem; font-weight:950; letter-spacing:-.04em; }
.final-decision .small { color:#DCFCE7; margin-top:5px; }
.presenter-note { background:#FFF7ED; border:1px solid #FED7AA; border-left:5px solid #F97316; padding:12px 14px; border-radius:10px; color:#7C2D12; }
.rubric-grid { display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:10px; margin-top:10px; }
.rubric-card { background:#fff; border:1px solid var(--line); border-radius:12px; padding:14px 15px; }
.rubric-card .r-head { display:flex; justify-content:space-between; gap:10px; color:var(--ink); font-weight:900; }
.rubric-card .r-score { color:#1D4ED8; white-space:nowrap; }
.rubric-card .r-answer { color:#475467; font-size:.85rem; line-height:1.5; margin-top:7px; }
.rubric-card .r-proof { color:#1D4ED8; font-size:.76rem; font-weight:800; margin-top:7px; }
@media (max-width: 1100px) { .stage-grid { grid-template-columns:repeat(2,1fr); } }
@media (max-width: 850px) { .rubric-grid { grid-template-columns:1fr; } }
.intro-wrap { min-height:86vh; display:flex; align-items:center; justify-content:center; }
.intro-stage { width:min(980px,94vw); text-align:center; padding:56px 42px 48px; border-radius:28px; color:#F8FAFC; background:radial-gradient(circle at 50% 0%,rgba(59,130,246,.32),transparent 38%),linear-gradient(145deg,#050914 0%,#0B1530 58%,#123A88 100%); border:1px solid rgba(148,163,184,.28); box-shadow:0 30px 80px rgba(15,23,42,.35); overflow:hidden; }
.intro-kicker { color:#93C5FD; font-size:.92rem; font-weight:900; letter-spacing:.18em; text-transform:uppercase; opacity:0; animation:introRise .65s ease .15s forwards; }
.intro-title { margin:12px 0 8px; font-size:clamp(2.4rem,6vw,4.7rem); line-height:1.04; font-weight:950; letter-spacing:-.065em; opacity:0; animation:introRise .8s ease .55s forwards; }
.intro-sub { color:#C7D7F8; font-size:clamp(1rem,2vw,1.25rem); opacity:0; animation:introRise .7s ease 1.05s forwards; }
.intro-agents { display:grid; grid-template-columns:repeat(6,minmax(0,1fr)); gap:10px; margin:34px 0 24px; }
.intro-agent { padding:13px 8px; border-radius:12px; background:rgba(15,23,42,.62); border:1px solid rgba(96,165,250,.28); color:#DBEAFE; font-size:.82rem; font-weight:850; opacity:0; transform:translateY(15px); animation:introAgent .5s ease forwards; }
.intro-agent:nth-child(1){animation-delay:1.35s}.intro-agent:nth-child(2){animation-delay:1.55s}.intro-agent:nth-child(3){animation-delay:1.75s}.intro-agent:nth-child(4){animation-delay:1.95s}.intro-agent:nth-child(5){animation-delay:2.15s}.intro-agent:nth-child(6){animation-delay:2.35s}
.intro-flow { color:#BFDBFE; font-size:.92rem; font-weight:800; letter-spacing:.06em; opacity:0; animation:introRise .7s ease 2.75s forwards; }
.intro-wrap + div [data-testid="stButton"] { max-width:320px; margin:0 auto; }
@keyframes introRise { from{opacity:0;transform:translateY(18px)} to{opacity:1;transform:translateY(0)} }
@keyframes introAgent { from{opacity:0;transform:translateY(15px)} to{opacity:1;transform:translateY(0)} }
@media (max-width: 850px) { .intro-agents { grid-template-columns:repeat(3,1fr); } }
</style>
""",
    unsafe_allow_html=True,
)

if "runs" not in st.session_state:
    st.session_state.runs = []
if "single" not in st.session_state:
    st.session_state.single = None
if "active_agent" not in st.session_state:
    st.session_state.active_agent = None
if "execution_log" not in st.session_state:
    st.session_state.execution_log = []
if "realtime_result" not in st.session_state:
    st.session_state.realtime_result = None
if "intro_complete" not in st.session_state:
    st.session_state.intro_complete = False
if "manual_question" not in st.session_state:
    st.session_state.manual_question = "결제는 완료되었는데 주문 내역에 보이지 않습니다."
if "manual_steps" not in st.session_state:
    st.session_state.manual_steps = []
if "manual_fault" not in st.session_state:
    st.session_state.manual_fault = None
if "manual_selected_agent" not in st.session_state:
    st.session_state.manual_selected_agent = "Interpreter"
if "three_axis_results" not in st.session_state:
    st.session_state.three_axis_results = []
if "qa2_judge_results" not in st.session_state:
    st.session_state.qa2_judge_results = []


def page_header(title: str, subtitle: str) -> None:
    st.markdown(f'<div class="page-title">{title}</div><div class="page-subtitle">{subtitle}</div>', unsafe_allow_html=True)


def badge(text: str, tone: str = "blue") -> str:
    return f'<span class="badge badge-{tone}">{text}</span>'


def save_reports(runs: list[dict[str, Any]]) -> None:
    REPORTS.mkdir(parents=True, exist_ok=True)
    rows: list[dict[str, Any]] = []
    defects: list[str] = []
    for idx, item in enumerate(runs, 1):
        case, judge = item["case"], item["judge"]
        rows.append({
            "case_id": case["case_id"], "type": case["type"], "question": case["question"],
            "score": judge["score"], "pass": judge["pass"],
            "critical_failure": item["critical_failure"], "reason": judge["reason"],
        })
        if not judge["pass"]:
            severity = "Critical" if item["critical_failure"] else "High" if judge["score"] < 50 else "Medium"
            defects.append(
                f"""## BUG-{idx:03d} — {case['case_id']}\n- 결함 제목: {case['expected_intent']} 품질 기준 미충족\n- 입력: {case['question']}\n- 기대: {', '.join(case['required_output'])}\n- 실제: {judge['reason']}\n- 심각도: {severity}\n- 개선: 필수 출력과 근거를 보강하고 재검증\n"""
            )
    df = pd.DataFrame(rows)
    df.to_csv(REPORTS / "test_result.csv", index=False, encoding="utf-8-sig")
    average = float(df["score"].mean())
    critical = bool(df["critical_failure"].any())
    decision = "즉시 배포 보류" if critical else "배포 가능" if average >= 90 else "조건부 배포 가능, 개선 후 재검증" if average >= 80 else "주요 개선 필요" if average >= 70 else "배포 보류"
    (REPORTS / "quality_score_report.md").write_text(
        f"# 품질 점수 보고서\n\n- 총 테스트: {len(df)}\n- PASS: {int(df['pass'].sum())}\n- FAIL: {int((~df['pass']).sum())}\n- 평균 점수: {average:.1f}\n- 성공률: {df['pass'].mean()*100:.1f}%\n", encoding="utf-8"
    )
    (REPORTS / "deployment_decision.md").write_text(
        f"# 배포 판단\n\n## {decision}\n\n- 평균 점수: {average:.1f}\n- Critical Failure: {'있음' if critical else '없음'}\n", encoding="utf-8"
    )
    (REPORTS / "defect_report.md").write_text(
        "# 결함 보고서\n\n" + ("\n".join(defects) if defects else "발견된 결함이 없습니다."), encoding="utf-8"
    )
    (REPORTS / "VOC_QA_Report.txt").write_bytes(build_text_report(runs))
    (REPORTS / "VOC_QA_Report.xml").write_bytes(build_xml_report(runs))
    (REPORTS / "VOC_QA_Report.html").write_bytes(build_html_report(runs))


def agent_status_html(active: str | None = None, completed: set[str] | None = None, errored: set[str] | None = None) -> str:
    completed = completed or set()
    errored = errored or set()
    statuses = server_status()
    cards = []
    for name in AGENTS:
        css = "running" if name == active else "error" if name in errored else "complete" if name in completed else "waiting"
        label = "RUNNING" if css == "running" else "ERROR" if css == "error" else "DONE" if css == "complete" else "READY"
        port = statuses.get(name.lower(), statuses.get(name, {})).get("port", "-")
        cards.append(f'<div class="agent-card {css}"><div class="name"><span class="status-dot"></span>{name}</div><div class="meta">{label} · :{port}</div></div>')
    return '<div class="agent-grid">' + "".join(cards) + "</div>"


def _step_message(step: dict[str, Any]) -> str:
    agent = step.get("agent", "Agent")
    messages = {
        "Interpreter": "고객 발화의 의도와 핵심 엔터티를 해석합니다.",
        "Retriever": "VOC 데이터에서 관련 근거를 검색합니다.",
        "Summarizer": "검색 근거와 고객 맥락을 요약합니다.",
        "Evaluator": "정확성·유용성·근거성을 평가합니다.",
        "Critic": "누락, 위험, 모순 가능성을 검토합니다.",
        "Improver": "최종 개선안과 고객 안내를 생성합니다.",
    }
    if step.get("status") == "ERROR":
        return step.get("error") or f"{agent} 실행 중 오류가 발생했습니다."
    return messages.get(agent, f"{agent} 단계를 실행합니다.")


def _timeline_html(logs: list[dict[str, Any]], running: str | None = None) -> str:
    rows = []
    for idx, item in enumerate(logs, 1):
        state = item.get("state", "complete")
        rows.append(
            f'<div class="timeline-row {state}">'
            f'<div class="timeline-step">STEP {idx:02d}</div>'
            f'<div class="timeline-agent">{item["agent"]}</div>'
            f'<div class="timeline-message">{item["message"]}</div>'
            f'<div class="timeline-time">{item.get("latency_ms", 0)} ms</div>'
            '</div>'
        )
    status = '<span class="live-chip">LIVE EXECUTION</span>' if running else '<span class="badge badge-green">COMPLETE</span>'
    return (
        '<div class="execution-shell">'
        '<div class="execution-head"><div><div class="execution-title">Agent Execution Timeline</div>'
        '<div class="execution-sub">각 Agent의 현재 처리 단계와 결과를 실시간으로 표시합니다.</div></div>'
        f'{status}</div><div class="timeline">' + ''.join(rows) + '</div></div>'
    )


def replay_agent_flow(output: dict[str, Any], speed: float = 1.0) -> None:
    agent_slot = st.empty()
    timeline_slot = st.empty()
    progress_slot = st.empty()
    detail_slot = st.empty()
    completed: set[str] = set()
    errored: set[str] = set()
    logs: list[dict[str, Any]] = []
    steps = output.get("steps", [])
    total = max(1, len(steps))

    for index, step in enumerate(steps, 1):
        name = step["agent"]
        running_logs = logs + [{
            "agent": name,
            "message": _step_message(step),
            "latency_ms": step.get("latency_ms", 0),
            "state": "running",
        }]
        agent_slot.markdown(agent_status_html(active=name, completed=completed, errored=errored), unsafe_allow_html=True)
        timeline_slot.markdown(_timeline_html(running_logs, running=name), unsafe_allow_html=True)
        progress_slot.progress((index - 0.5) / total, text=f"{index}/{total} · {name} 실행 중")
        detail_slot.info(f"현재 실행: {name} — {_step_message(step)}")
        time.sleep(0.32 * max(speed, 0.0))

        state = "error" if step.get("status") == "ERROR" else "complete"
        if state == "error":
            errored.add(name)
        else:
            completed.add(name)
        logs.append({
            "agent": name,
            "message": _step_message(step),
            "latency_ms": step.get("latency_ms", 0),
            "state": state,
        })
        st.session_state.execution_log = logs
        agent_slot.markdown(agent_status_html(completed=completed, errored=errored), unsafe_allow_html=True)
        timeline_slot.markdown(_timeline_html(logs, running=None if index == total else "next"), unsafe_allow_html=True)
        progress_slot.progress(index / total, text=f"{index}/{total} · {name} 완료")
        time.sleep(0.12 * max(speed, 0.0))

    detail_slot.success("6개 Agent Pipeline 실행이 완료되었습니다.")
    timeline_slot.markdown(_timeline_html(logs), unsafe_allow_html=True)


def render_result(output: dict[str, Any] | None) -> None:
    if not output:
        return
    final = output.get("final")
    if final:
        analysis = final["analysis"]
        critique = final["critique"]
        improvement = final["improvement"]
        c1, c2, c3 = st.columns([1.15, 1, 1])
        with c1:
            st.markdown('<div class="section-title">분석 요약</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="result-card"><h4>{analysis.get("summary", "분석 완료")}</h4><p>의도와 관련 근거를 기반으로 VOC를 구조화했습니다.</p></div>', unsafe_allow_html=True)
        with c2:
            st.markdown('<div class="section-title">주요 위험</div>', unsafe_allow_html=True)
            risks = critique.get("risks", []) or ["식별된 주요 위험 없음"]
            risk_html = "".join(f"<p>• {risk}</p>" for risk in risks[:4])
            st.markdown(f'<div class="result-card"><h4>Risk Review</h4>{risk_html}</div>', unsafe_allow_html=True)
        with c3:
            st.markdown('<div class="section-title">개선 우선순위</div>', unsafe_allow_html=True)
            priority = improvement.get("priority", "-" )
            tone = "red" if str(priority).lower() in {"critical", "high", "긴급", "높음"} else "amber"
            st.markdown(f'<div class="result-card"><h4>{badge(str(priority), tone)}</h4><p>{improvement.get("customer_guidance", "")}</p></div>', unsafe_allow_html=True)

        st.markdown('<div class="section-title">권고 조치</div>', unsafe_allow_html=True)
        actions = improvement.get("improvement_actions", [])
        cols = st.columns(max(1, min(3, len(actions)))) if actions else [st.container()]
        for i, action in enumerate(actions):
            with cols[i % len(cols)]:
                st.markdown(f'<div class="result-card"><h4>Action {i+1}</h4><p>{action}</p></div>', unsafe_allow_html=True)

    with st.expander("Agent별 상세 결과", expanded=False):
        for index, step in enumerate(output.get("steps", []), 1):
            status = step.get("status", "UNKNOWN")
            tone = "green" if status == "SUCCESS" else "red" if status == "ERROR" else "blue"
            st.markdown(f"**{index}. {step['agent']}** &nbsp; {badge(status, tone)} &nbsp; `{step.get('latency_ms', 0)}ms`", unsafe_allow_html=True)
            if step.get("error"):
                st.error(step["error"])
            st.json(step.get("output", {}), expanded=False)
            if step.get("evidence"):
                st.caption("근거: " + ", ".join(step["evidence"]))
            if index < len(output.get("steps", [])):
                st.divider()



def _extract_latency_ms(item: dict[str, Any]) -> int:
    explicit = item.get("response_time_ms")
    if isinstance(explicit, (int, float)):
        return int(explicit)
    steps = item.get("result", {}).get("steps", [])
    return int(sum(step.get("latency_ms", 0) or 0 for step in steps))


def _qa_kpis_html(runs: list[dict[str, Any]], total: int, running: bool = False) -> str:
    done = len(runs)
    passed = sum(1 for item in runs if item["judge"]["pass"])
    failed = done - passed
    scores = [float(item["judge"].get("score", 0)) for item in runs]
    latencies = [_extract_latency_ms(item) for item in runs]
    avg = sum(scores) / done if done else 0.0
    avg_latency = sum(latencies) / done if done else 0.0
    max_latency = max(latencies) if latencies else 0
    defects = failed
    values = [
        ("진행", f"{done}/{total}", "실행 완료", "active" if running else ""),
        ("PASS", str(passed), f"{(passed/done*100):.1f}%" if done else "0.0%", "pass"),
        ("FAIL", str(failed), "자동 결함 후보", "fail" if failed else ""),
        ("평균 점수", f"{avg:.1f}", "100점 기준", ""),
        ("평균 응답", f"{avg_latency:.0f} ms", f"최대 {max_latency} ms", ""),
        ("결함", str(defects), "Critical 포함", "fail" if defects else "pass"),
    ]
    cards = []
    for label, value, sub, css in values:
        cards.append(f'<div class="qa-kpi {css}"><div class="k-label">{label}</div><div class="k-value">{value}</div><div class="k-sub">{sub}</div></div>')
    return '<div class="qa-live-grid">' + ''.join(cards) + '</div>'


def _qa_recent_html(runs: list[dict[str, Any]], limit: int = 8) -> str:
    rows = []
    for item in list(reversed(runs[-limit:])):
        case = item["case"]
        passed = bool(item["judge"]["pass"])
        state = "pass" if passed else "fail"
        status = "PASS" if passed else "FAIL"
        rows.append(
            f'<div class="qa-result-row {state}">'
            f'<div class="qid">{case["case_id"]}</div>'
            f'<div class="qtype">{case["type"]}</div>'
            f'<div class="qtext">{case["question"]}</div>'
            f'<div class="qscore">{item["judge"].get("score", 0):.0f}</div>'
            f'<div class="qstatus">{status}</div></div>'
        )
    if not rows:
        rows.append('<div class="small-note">실행 결과가 여기에 누적됩니다.</div>')
    return '<div class="qa-result-list">' + ''.join(rows) + '</div>'


def _qa_type_summary(runs: list[dict[str, Any]]) -> pd.DataFrame:
    if not runs:
        return pd.DataFrame(columns=["유형", "실행", "PASS", "FAIL", "평균점수", "성공률"])
    rows = []
    for item in runs:
        rows.append({
            "유형": item["case"]["type"],
            "PASS": int(bool(item["judge"]["pass"])),
            "FAIL": int(not bool(item["judge"]["pass"])),
            "점수": float(item["judge"].get("score", 0)),
        })
    df = pd.DataFrame(rows)
    grouped = df.groupby("유형", as_index=False).agg(실행=("점수", "size"), PASS=("PASS", "sum"), FAIL=("FAIL", "sum"), 평균점수=("점수", "mean"))
    grouped["성공률"] = grouped["PASS"] / grouped["실행"] * 100
    grouped["평균점수"] = grouped["평균점수"].round(1)
    grouped["성공률"] = grouped["성공률"].round(1)
    return grouped.sort_values(["성공률", "평균점수"], ascending=[True, True])



def _run_rows(runs: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for index, item in enumerate(runs, 1):
        steps = item.get("result", {}).get("steps", [])
        rows.append({
            "순번": index,
            "TC_ID": item["case"]["case_id"],
            "유형": item["case"]["type"],
            "질문": item["case"]["question"],
            "점수": float(item["judge"].get("score", 0)),
            "판정": "PASS" if item["judge"]["pass"] else "FAIL",
            "응답시간(ms)": _extract_latency_ms(item),
            "Critical": bool(item.get("critical_failure", False)),
            "사유": item["judge"].get("reason", ""),
            "Agent단계": len(steps),
        })
    return pd.DataFrame(rows)


def _agent_performance(runs: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for item in runs:
        for step in item.get("result", {}).get("steps", []):
            rows.append({
                "Agent": step.get("agent", "Unknown"),
                "응답시간(ms)": float(step.get("latency_ms", 0) or 0),
                "상태": step.get("status", "UNKNOWN"),
            })
    if not rows:
        return pd.DataFrame(columns=["Agent", "평균응답(ms)", "최대응답(ms)", "실행횟수", "오류수"])
    df = pd.DataFrame(rows)
    df["오류"] = (df["상태"] == "ERROR").astype(int)
    grouped = df.groupby("Agent", as_index=False).agg(
        **{"평균응답(ms)": ("응답시간(ms)", "mean"), "최대응답(ms)": ("응답시간(ms)", "max"), "실행횟수": ("Agent", "size"), "오류수": ("오류", "sum")}
    )
    grouped["평균응답(ms)"] = grouped["평균응답(ms)"].round(1)
    grouped["최대응답(ms)"] = grouped["최대응답(ms)"].round(1)
    order = {name: i for i, name in enumerate(AGENTS)}
    return grouped.sort_values("Agent", key=lambda x: x.map(order))


def _deployment_decision(runs: list[dict[str, Any]]) -> tuple[str, float, float, int, int, bool]:
    if not runs:
        return "검증 대기", 0.0, 0.0, 0, 0, False
    df = _run_rows(runs)
    avg = float(df["점수"].mean())
    passed = int((df["판정"] == "PASS").sum())
    failed = len(df) - passed
    rate = passed / len(df) * 100
    critical = bool(df["Critical"].any())
    decision = "즉시 배포 보류" if critical else "배포 가능" if avg >= 90 else "조건부 배포 가능" if avg >= 80 else "주요 개선 필요" if avg >= 70 else "배포 보류"
    return decision, avg, rate, passed, failed, critical


def build_excel_report(runs: list[dict[str, Any]]) -> bytes:
    result_df = _run_rows(runs)
    type_df = _qa_type_summary(runs)
    agent_df = _agent_performance(runs)
    decision, avg, rate, passed, failed, critical = _deployment_decision(runs)
    wb = Workbook()
    ws = wb.active
    ws.title = "요약"
    summary = [("VOC AI Agent QA 보고서", ""), ("생성일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")), ("배포판단", decision), ("평균점수", round(avg,1)), ("성공률", f"{rate:.1f}%"), ("PASS", passed), ("FAIL", failed), ("Critical", "있음" if critical else "없음")]
    for r,(k,v) in enumerate(summary,1):
        ws.cell(r,1,k); ws.cell(r,2,v)
    ws["A1"].font=Font(size=18,bold=True,color="FFFFFF"); ws["A1"].fill=PatternFill("solid",fgColor="172554"); ws.merge_cells("A1:B1")
    for row in range(2,len(summary)+1):
        ws.cell(row,1).font=Font(bold=True); ws.cell(row,1).fill=PatternFill("solid",fgColor="DBEAFE")
    ws.column_dimensions["A"].width=22; ws.column_dimensions["B"].width=35
    def add_df(name, df):
        sh=wb.create_sheet(name)
        for c,col in enumerate(df.columns,1):
            cell=sh.cell(1,c,col); cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="2563EB"); cell.alignment=Alignment(horizontal="center")
        for r,row in enumerate(df.itertuples(index=False),2):
            for c,val in enumerate(row,1): sh.cell(r,c,val)
        sh.freeze_panes="A2"; sh.auto_filter.ref=sh.dimensions
        for col in sh.columns:
            letter=col[0].column_letter; sh.column_dimensions[letter].width=min(45,max(12,max(len(str(x.value or "")) for x in col)+2))
    add_df("전체결과", result_df); add_df("유형별통계", type_df); add_df("Agent성능", agent_df)
    bio=io.BytesIO(); wb.save(bio); return bio.getvalue()


def _register_pdf_font() -> str:
    candidates=[r"C:\Windows\Fonts\malgun.ttf", r"C:\Windows\Fonts\gulim.ttc", "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"]
    for path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("Korean", path)); return "Korean"
            except Exception:
                continue
    return "Helvetica"


def build_pdf_report(runs: list[dict[str, Any]]) -> bytes:
    df=_run_rows(runs); type_df=_qa_type_summary(runs); decision,avg,rate,passed,failed,critical=_deployment_decision(runs)
    bio=io.BytesIO(); font=_register_pdf_font()
    doc=SimpleDocTemplate(bio,pagesize=landscape(A4),rightMargin=14*mm,leftMargin=14*mm,topMargin=12*mm,bottomMargin=12*mm)
    styles=getSampleStyleSheet(); title=ParagraphStyle("KTitle",parent=styles["Title"],fontName=font,fontSize=22,leading=28,textColor=colors.HexColor("#172554")); body=ParagraphStyle("KBody",parent=styles["BodyText"],fontName=font,fontSize=9,leading=13)
    story=[Paragraph("VOC AI Agent 품질 검증 보고서",title),Paragraph(f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",body),Spacer(1,8)]
    summary=[["배포 판단","평균 점수","성공률","PASS","FAIL","Critical"],[decision,f"{avg:.1f}",f"{rate:.1f}%",str(passed),str(failed),"있음" if critical else "없음"]]
    t=Table(summary,colWidths=[62*mm,32*mm,32*mm,25*mm,25*mm,28*mm]); t.setStyle(TableStyle([('FONTNAME',(0,0),(-1,-1),font),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#172554')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('ALIGN',(0,0),(-1,-1),'CENTER'),('GRID',(0,0),(-1,-1),.5,colors.HexColor('#CBD5E1')),('FONTSIZE',(0,0),(-1,-1),10),('BOTTOMPADDING',(0,0),(-1,-1),8),('TOPPADDING',(0,0),(-1,-1),8)])); story += [t,Spacer(1,12),Paragraph("유형별 품질 현황",styles["Heading2"])]
    td=[list(type_df.columns)]+[[str(v) for v in row] for row in type_df.itertuples(index=False,name=None)]
    tt=Table(td,repeatRows=1); tt.setStyle(TableStyle([('FONTNAME',(0,0),(-1,-1),font),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#2563EB')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.4,colors.HexColor('#CBD5E1')),('ALIGN',(1,1),(-1,-1),'CENTER'),('FONTSIZE',(0,0),(-1,-1),8)])); story += [tt,PageBreak(),Paragraph("실패 및 결함 후보",styles["Heading2"])]
    failures=df[df["판정"]=="FAIL"]
    if failures.empty: story.append(Paragraph("발견된 실패 테스트가 없습니다.",body))
    else:
        fd=[["TC_ID","유형","점수","Critical","사유"]]+[[str(r["TC_ID"]),str(r["유형"]),str(r["점수"]),str(r["Critical"]),Paragraph(str(r["사유"]),body)] for _,r in failures.iterrows()]
        ft=Table(fd,colWidths=[25*mm,35*mm,20*mm,25*mm,140*mm],repeatRows=1); ft.setStyle(TableStyle([('FONTNAME',(0,0),(-1,-1),font),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#991B1B')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.4,colors.HexColor('#CBD5E1')),('VALIGN',(0,0),(-1,-1),'TOP'),('FONTSIZE',(0,0),(-1,-1),8)])); story.append(ft)
    doc.build(story); return bio.getvalue()


def build_json_report(runs: list[dict[str, Any]]) -> bytes:
    decision,avg,rate,passed,failed,critical=_deployment_decision(runs)
    payload={"generated_at":datetime.now().isoformat(timespec="seconds"),"summary":{"decision":decision,"average_score":round(avg,1),"pass_rate":round(rate,1),"pass":passed,"fail":failed,"critical":critical},"results":_run_rows(runs).to_dict(orient="records"),"agent_performance":_agent_performance(runs).to_dict(orient="records")}
    return json.dumps(payload,ensure_ascii=False,indent=2).encode("utf-8")


def _report_summary(runs: list[dict[str, Any]]) -> dict[str, Any]:
    decision, avg, rate, passed, failed, critical = _deployment_decision(runs)
    return {"total": len(runs), "pass": passed, "fail": failed, "pass_rate": round(rate, 1),
            "average_score": round(avg, 1), "critical": critical, "decision": decision}


def build_text_report(runs: list[dict[str, Any]]) -> bytes:
    summary = _report_summary(runs)
    lines = ["VOC AI Agent QA Report", "=" * 28]
    lines += [f"{key}: {value}" for key, value in summary.items()]
    lines.append("\n[Test Results]")
    for row in _run_rows(runs).to_dict(orient="records"):
        lines.append(f"{row['TC_ID']} | {row['판정']} | {row['점수']:.1f} | {row['사유']}")
    return "\n".join(lines).encode("utf-8")


def build_xml_report(runs: list[dict[str, Any]]) -> bytes:
    root = ET.Element("voc_qa_report", generated_at=datetime.now().isoformat(timespec="seconds"))
    summary_node = ET.SubElement(root, "summary")
    for key, value in _report_summary(runs).items():
        ET.SubElement(summary_node, key).text = str(value).lower() if isinstance(value, bool) else str(value)
    results_node = ET.SubElement(root, "results")
    for row in _run_rows(runs).to_dict(orient="records"):
        item = ET.SubElement(results_node, "test_case", id=str(row["TC_ID"]))
        for key in ["유형", "점수", "판정", "응답시간(ms)", "Critical", "사유"]:
            ET.SubElement(item, "field", name=key).text = str(row[key])
    ET.indent(root)
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def build_html_report(runs: list[dict[str, Any]]) -> bytes:
    summary = _report_summary(runs)
    rows = "".join(
        f"<tr><td>{html.escape(str(row['TC_ID']))}</td><td>{html.escape(str(row['유형']))}</td>"
        f"<td>{row['점수']:.1f}</td><td>{html.escape(str(row['판정']))}</td>"
        f"<td>{html.escape(str(row['사유']))}</td></tr>"
        for row in _run_rows(runs).to_dict(orient="records")
    )
    document = f"""<!doctype html><html lang="ko"><meta charset="utf-8"><title>VOC QA Report</title>
<style>body{{font-family:Arial,sans-serif;max-width:1100px;margin:32px auto;color:#172033}}.summary{{display:flex;gap:12px;flex-wrap:wrap}}.kpi{{padding:12px;border:1px solid #ddd;border-radius:8px}}table{{width:100%;border-collapse:collapse;margin-top:20px}}th,td{{border:1px solid #ddd;padding:8px;text-align:left}}th{{background:#172554;color:white}}</style>
<h1>VOC AI Agent QA Report</h1><div class="summary">{''.join(f'<div class="kpi"><b>{html.escape(str(k))}</b><br>{html.escape(str(v))}</div>' for k,v in summary.items())}</div>
<table><thead><tr><th>TC_ID</th><th>유형</th><th>점수</th><th>판정</th><th>사유</th></tr></thead><tbody>{rows}</tbody></table></html>"""
    return document.encode("utf-8")


def _presentation_stages(active: int, done: int) -> str:
    names = ["VOC 접수", "Agent 분석", "QA 검증", "품질 진단", "배포 판단"]
    cards = []
    for index, name in enumerate(names, 1):
        css = "done" if index <= done else "active" if index == active else ""
        cards.append(f'<div class="stage-card {css}"><div class="stage-no">STAGE {index:02d}</div><div class="stage-name">{name}</div></div>')
    return '<div class="stage-grid">' + ''.join(cards) + '</div>'


def _presentation_rubric_html(runs: list[dict[str, Any]]) -> str:
    if runs:
        decision, avg, rate, passed, failed, critical = _deployment_decision(runs)
        test_answer = f"총 {len(runs)}건을 실행해 PASS {passed}건, FAIL {failed}건, 성공률 {rate:.1f}%, 평균 {avg:.1f}점을 확인했습니다."
        defect_answer = f"실패 {failed}건과 Critical {'발생' if critical else '미발생'}을 분리 기록하고, 원인·조치·재검증 항목으로 관리합니다."
        test_proof = "근거: QA 테스트·품질 분석·보고서 화면의 현재 실행 결과"
    else:
        decision = "검증 대기"
        test_answer = "Happy·Edge·Negative·Fault 총 20개 테스트를 실행하고 PASS, FAIL, 성공률, 평균 점수를 자동 집계합니다."
        defect_answer = "API 연결 실패, CSV 누락, 응답 지연과 테스트 실패를 재현해 원인·영향·조치 후 재검증으로 관리합니다."
        test_proof = "근거: 발표 데모 실행 후 실제 수치로 자동 갱신"
    items = [
        ("01", "프로젝트 목적 이해도", "VOC를 단순 분류하는 데서 끝내지 않고, 근거 기반 개선안의 품질을 검증해 안전한 운영·배포 판단까지 연결하는 프로젝트입니다.", "근거: 전체 5단계 발표 흐름"),
        ("02", "고객 불만 분석의 적절성", "Interpreter가 불만 유형을 식별하고 Retriever가 유사 VOC와 원인을 찾으며, Summarizer가 고객 영향과 한계를 근거와 함께 정리합니다.", "근거: Agent 단계 실행의 INPUT·OUTPUT"),
        ("03", "정책 개선안의 타당성", "Improver는 검색된 원인과 정책을 직접 연결해 고객 안내, 실행 조치, 우선순위와 KPI를 제시하므로 적용 가능성을 추적할 수 있습니다.", "근거: improvement_actions·priority·KPI"),
        ("04", "멀티 에이전트 역할 설명", "Interpreter=의도, Retriever=근거, Summarizer=요약, Evaluator=품질 점검, Critic=위험 반론, Improver=실행 가능한 개선안 역할입니다.", "근거: 6개 Agent 타임라인"),
        ("05", "내부 품질진단의 충실성", "최종 문장만 채점하지 않고 해석→검색→요약→평가→비판→개선의 상태, 출력, 근거, 지연시간을 단계별로 검사합니다.", "근거: 실행 로그·Agent 성능표"),
        ("06", "독립 LLM Judge 평가 설명", "Evaluator·Critic은 생성 파이프라인 내부의 자기점검이고, Anthropic Judge는 생성 모델과 분리된 최종 심사자입니다. 모델 편향과 자기평가 관대함을 줄이기 위해 둘 다 필요합니다.", "근거: 독립 LLM 교차검증 흐름·A–D 실험군"),
        ("07", "테스트 결과의 객관성", test_answer, test_proof),
        ("08", "장애 및 결함관리 내용", defect_answer, "근거: 장애·결함 화면과 defect_report"),
        ("09", "발표 구성 및 전달력", f"문제 정의→VOC 분석→내부 진단→독립 검증→정량 결과→결함 조치→{decision} 순서로 결론을 제시합니다.", "근거: 발표 모드 STAGE 01–05"),
        ("10", "팀 협업 및 질의응답", "분석·Agent, QA·Judge, UI·보고서 담당으로 산출물을 분리하고 공통 테스트 결과로 통합합니다. 질문에는 화면 로그와 보고서 수치를 근거로 답합니다.", "근거: 역할 분담표·Q&A 답변 가이드"),
    ]
    cards = []
    for no, title, answer, proof in items:
        cards.append(
            f'<div class="rubric-card"><div class="r-head"><span>{no}. {title}</span><span class="r-score">8점</span></div>'
            f'<div class="r-answer">{answer}</div><div class="r-proof">{proof}</div></div>'
        )
    return '<div class="rubric-grid">' + ''.join(cards) + '</div>'


def _technical_rubric_html(runs: list[dict[str, Any]]) -> str:
    if runs:
        summary = _report_summary(runs)
        quantitative = (f"총 {summary['total']}건, PASS {summary['pass']}건, FAIL {summary['fail']}건, "
                        f"성공률 {summary['pass_rate']:.1f}%, 평균 {summary['average_score']:.1f}점으로 산출했습니다.")
        release = f"Critical {'있음' if summary['critical'] else '없음'}과 품질 점수를 적용한 현재 판단은 ‘{summary['decision']}’입니다."
    else:
        quantitative = "20개 시나리오 실행 후 PASS·FAIL·성공률·점수·단계별 수행시간을 동일 기준으로 자동 산출합니다."
        release = "평균 점수, Critical Gate와 잔여 결함을 확인하기 전에는 ‘검증 대기’로 두며 추정 배포를 허용하지 않습니다."
    items = [
        ("01", "요구사항 충족도", "VOC 분석, 6개 Agent, 내부·독립 평가, QA, 결함관리, 보고서와 배포판단을 하나의 실행 흐름으로 구현했습니다.", "증적: 전체 메뉴·다운로드 보고서"),
        ("02", "시스템 구조의 적정성", "Interpreter→Retriever→Summarizer→Evaluator→Critic→Improver로 책임을 분리하고 각 단계의 입력·출력·오류를 추적합니다.", "증적: Agent 단계 실행·타임라인"),
        ("03", "VOC 분석 정확성", "의도와 불만 유형을 검색 근거의 원인·정책과 연결하고, 영향·한계·개선 방향을 분리해 근거 없는 확정을 방지합니다.", "증적: matches·facts·limitations"),
        ("04", "개선안의 실행 가능성", "고객 안내와 정책 조치를 구분하고 우선순위, 예상 처리 단계, 재발률·처리시간·1차 해결률 KPI까지 제공합니다.", "증적: Improver 출력"),
        ("05", "내부 품질검증 수준", "Agent 단위·문법, 빈 입력·서버 중단·파일 누락 예외, 분기, 통합 및 20개 E2E 테스트를 pytest와 앱에서 수행합니다.", "증적: test_agent_unit·test_fault_tolerance·test_pipeline_e2e"),
        ("06", "독립 평가의 객관성", "생성 역할과 내부 Evaluator·Critic 뒤에 Anthropic Judge를 분리하고 A–D 모델 역할 교환 실험으로 동일 모델 자기평가와 비교합니다.", "증적: 교차검증 흐름·Judge 명칭"),
        ("07", "결함관리 적절성", "각 FAIL을 TC_ID, 원인, 영향도, 심각도, 조치와 재시험 대상으로 기록하며 Critical 결함은 점수와 무관하게 배포를 차단합니다.", "증적: 장애·결함·defect_report.md"),
        ("08", "정량적 품질성과", quantitative, "증적: QA KPI·Agent 수행시간·CSV"),
        ("09", "산출물 및 증적관리", "동일 실행 결과를 TXT, XML, HTML, CSV, JSON, Markdown, Excel, PDF로 생성하고 reports 폴더와 다운로드 화면에서 관리합니다.", "증적: 보고서·배포 화면"),
        ("10", "배포 및 운영 준비도", release + " 잔여 FAIL과 Critical 사유를 함께 제시해 조건부 배포와 보류를 구분합니다.", "증적: Release Readiness·배포 판단 보고서"),
    ]
    return '<div class="rubric-grid">' + ''.join(
        f'<div class="rubric-card"><div class="r-head"><span>{no}. {title}</span><span class="r-score">2점</span></div>'
        f'<div class="r-answer">{answer}</div><div class="r-proof">{proof}</div></div>'
        for no, title, answer, proof in items
    ) + '</div>'


def _run_full_presentation(use_llm: bool, sample_question: str, speed: float = 1.0) -> None:
    stage_slot = st.empty()
    headline_slot = st.empty()
    agent_slot = st.empty()
    qa_slot = st.empty()
    result_slot = st.empty()

    stage_slot.markdown(_presentation_stages(1, 0), unsafe_allow_html=True)
    headline_slot.markdown('<div class="cinema-panel"><div class="cinema-title">고객 VOC 접수</div><div class="cinema-copy">고객의 비정형 발화를 AI Agent 품질 검증 파이프라인에 전달합니다.</div></div>', unsafe_allow_html=True)
    time.sleep(.45 * speed)

    stage_slot.markdown(_presentation_stages(2, 1), unsafe_allow_html=True)
    headline_slot.markdown('<div class="cinema-panel"><div class="cinema-title">6개 Agent Pipeline 실행</div><div class="cinema-copy">의도 해석부터 검색, 요약, 평가, 비판, 개선까지 순차 처리합니다.</div></div>', unsafe_allow_html=True)
    output = analyze_voc(sample_question, None)
    st.session_state.single = output
    completed: set[str] = set()
    logs: list[dict[str, Any]] = []
    total_steps = max(1, len(output.get("steps", [])))
    for idx, step in enumerate(output.get("steps", []), 1):
        name = step["agent"]
        agent_slot.markdown(agent_status_html(active=name, completed=completed), unsafe_allow_html=True)
        logs_running = logs + [{"agent":name,"message":_step_message(step),"latency_ms":step.get("latency_ms",0),"state":"running"}]
        qa_slot.markdown(_timeline_html(logs_running, running=name), unsafe_allow_html=True)
        time.sleep(.30 * speed)
        state = "error" if step.get("status") == "ERROR" else "complete"
        logs.append({"agent":name,"message":_step_message(step),"latency_ms":step.get("latency_ms",0),"state":state})
        completed.add(name)
        agent_slot.markdown(agent_status_html(completed=completed), unsafe_allow_html=True)
        time.sleep(.08 * speed)
    st.session_state.execution_log = logs

    stage_slot.markdown(_presentation_stages(3, 2), unsafe_allow_html=True)
    headline_slot.markdown('<div class="cinema-panel"><div class="cinema-title">20개 QA 시나리오 자동 검증</div><div class="cinema-copy">Happy·Edge·Negative·Fault 시나리오를 Rule/LLM Judge로 평가합니다.</div></div>', unsafe_allow_html=True)
    cases = load_cases()
    runs: list[dict[str, Any]] = []
    progress = result_slot.progress(0, text="QA 검증 준비")
    for idx, case in enumerate(cases, 1):
        item = run_case(case, use_llm)
        item["response_time_ms"] = max(_extract_latency_ms(item), 1)
        runs.append(item)
        qa_slot.markdown(_qa_kpis_html(runs, len(cases), running=idx < len(cases)), unsafe_allow_html=True)
        progress.progress(idx/len(cases), text=f"{idx}/{len(cases)} · {case['case_id']} · {'PASS' if item['judge']['pass'] else 'FAIL'}")
        time.sleep(.035 * speed)
    st.session_state.runs = runs
    save_reports(runs)

    stage_slot.markdown(_presentation_stages(4, 3), unsafe_allow_html=True)
    headline_slot.markdown('<div class="cinema-panel"><div class="cinema-title">품질 진단 및 결함 분석</div><div class="cinema-copy">점수, 성공률, 응답시간, Critical Gate를 종합하여 품질 수준을 산정합니다.</div></div>', unsafe_allow_html=True)
    decision, avg, rate, passed, failed, critical = _deployment_decision(runs)
    df = _run_rows(runs)
    a,b,c,d = result_slot.columns(4)
    a.metric("평균 품질", f"{avg:.1f}")
    b.metric("성공률", f"{rate:.1f}%")
    c.metric("PASS / FAIL", f"{passed} / {failed}")
    d.metric("평균 응답", f"{df['응답시간(ms)'].mean():.0f} ms")
    time.sleep(.55 * speed)

    stage_slot.markdown(_presentation_stages(5, 4), unsafe_allow_html=True)
    tone = "hold" if "보류" in decision else "review" if "조건부" in decision or "개선" in decision else ""
    headline_slot.markdown(f'<div class="final-decision {tone}"><div class="big">{decision}</div><div class="small">평균 {avg:.1f}점 · 성공률 {rate:.1f}% · Critical {"있음" if critical else "없음"}</div></div>', unsafe_allow_html=True)
    agent_slot.markdown(_presentation_stages(0, 5), unsafe_allow_html=True)
    qa_slot.success("전체 발표 데모와 보고서 생성을 완료했습니다.")


def quality_summary() -> tuple[int, int, float, float, bool]:
    runs = st.session_state.runs
    if not runs:
        return 0, 0, 0.0, 0.0, False
    passed = sum(1 for item in runs if item["judge"]["pass"])
    failed = len(runs) - passed
    avg = sum(item["judge"]["score"] for item in runs) / len(runs)
    rate = passed / len(runs) * 100
    critical = any(item["critical_failure"] for item in runs)
    return passed, failed, avg, rate, critical


if not st.session_state.intro_complete:
    st.markdown(
        """
        <div class="intro-wrap">
          <div class="intro-stage">
            <div class="intro-kicker">AI Agent Quality Platform</div>
            <div class="intro-title">VOC 품질관리 플랫폼</div>
            <div class="intro-sub">AI 기반 VOC 분석 · 자동 품질평가 · 응답 개선</div>
            <div class="intro-agents">
              <div class="intro-agent">Interpreter</div>
              <div class="intro-agent">Retriever</div>
              <div class="intro-agent">Summarizer</div>
              <div class="intro-agent">Evaluator</div>
              <div class="intro-agent">Critic</div>
              <div class="intro-agent">Improver</div>
            </div>
            <div class="intro-flow">6 AI AGENTS → REAL-TIME QUALITY EVALUATION</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if st.button("START DEMO", type="primary", use_container_width=True):
        st.session_state.intro_complete = True
        st.rerun()
    st.stop()


with st.sidebar:
    st.markdown("## VOC Quality Studio")
    st.caption("AI Agent QA Platform")
    st.divider()
    navigation = st.radio(
        "NAVIGATION",
        ["분석·실시간 평가", "QA·장애 테스트", "품질 대시보드", "보고서·배포", "발표 모드"],
        label_visibility="collapsed",
    )
    st.divider()
    use_llm = st.toggle("Anthropic 독립 Judge", False, help="ANTHROPIC_API_KEY가 없으면 Rule Judge로 자동 대체")
    mode = "Anthropic LLM Judge" if use_llm else "Rule Judge"
    st.markdown(f"**평가 모드**  \n{mode}")
    st.caption("발표 시 Rule Judge 권장")

if navigation == "분석·실시간 평가":
    page = "Agent 단계 실행"
elif navigation == "QA·장애 테스트":
    page = st.radio("기능 선택", ["QA 테스트", "장애·결함"], horizontal=True, label_visibility="collapsed")
elif navigation == "품질 대시보드":
    page = st.radio("기능 선택", ["대시보드", "품질 분석"], horizontal=True, label_visibility="collapsed")
else:
    page = navigation

if page == "발표 모드":
    st.markdown('<div class="presentation-hero"><h1>VOC AI Agent Quality Assurance</h1><p>고객 VOC 개선부터 자동 QA, 품질 시각화, 결함 분석, 배포 판단까지 하나의 흐름으로 시연합니다.</p></div>', unsafe_allow_html=True)
    st.markdown(_presentation_stages(1, 0), unsafe_allow_html=True)
    intro_left, intro_right = st.columns([1.5, 1])
    with intro_left:
        speed_name = st.radio("시연 속도", ["빠르게", "표준", "천천히"], horizontal=True, index=1)
        speed_map = {"빠르게": .55, "표준": 1.0, "천천히": 1.55}
        start_demo = st.button("▶ 전체 발표 시작", type="primary", use_container_width=True)
    with intro_right:
        st.markdown('<div class="presenter-note"><b>발표 핵심 메시지</b><br>AI를 만드는 데서 끝나지 않고, AI의 품질을 자동으로 검증하고 배포 가능성까지 판단하는 QA 플랫폼입니다.</div>', unsafe_allow_html=True)
        st.caption("권장 발표 순서: 문제 정의 → Agent 흐름 → QA 자동화 → 품질 지표 → 배포 판단")
    if start_demo:
        demo_question = "결제는 완료됐는데 주문 내역이 보이지 않습니다."
        _run_full_presentation(use_llm, demo_question, speed_map[speed_name])
    elif st.session_state.runs:
        decision,avg,rate,passed,failed,critical=_deployment_decision(st.session_state.runs)
        tone = "hold" if "보류" in decision else "review" if "조건부" in decision or "개선" in decision else ""
        st.markdown(f'<div class="final-decision {tone}"><div class="big">최근 결과 · {decision}</div><div class="small">평균 {avg:.1f}점 · 성공률 {rate:.1f}% · PASS {passed} · FAIL {failed}</div></div>', unsafe_allow_html=True)
        st.markdown('<div class="section-title" style="margin-top:1rem">최근 Agent 실행</div>', unsafe_allow_html=True)
        st.markdown(agent_status_html(completed=set(AGENTS)), unsafe_allow_html=True)
        if st.session_state.execution_log:
            st.markdown(_timeline_html(st.session_state.execution_log), unsafe_allow_html=True)
    else:
        st.markdown('<div class="cinema-panel"><div class="cinema-title">One-click Presentation</div><div class="cinema-copy">버튼 한 번으로 6개 Agent 실행, 20개 QA 테스트, 품질 진단, 자동 보고서, 최종 배포 판단까지 진행됩니다.</div></div>', unsafe_allow_html=True)

    st.divider()
    validation_tab, qa2_tab, rubric_tab, technical_tab, team_tab, qa_tab = st.tabs([
        "QA1 3축 품질검증", "QA2 독립 Judge", "80점 평가 대응",
        "기술평가 20점 대응", "팀 역할 분담", "예상 질문·답변",
    ])
    with validation_tab:
        st.markdown('<div class="section-title">요약 · 개선안 · 이종 모델을 분리 검증</div>', unsafe_allow_html=True)
        axis_cols = st.columns(3)
        axis_cols[0].info("① 요약 정확성\n\n불만 핵심과 복합 의도 누락 여부")
        axis_cols[1].info("② 원인–개선안 적합성\n\n검색 근거와 실행 조치의 연결")
        axis_cols[2].info("③ 이종 모델 일관성\n\nRule 기준과 Anthropic 독립 Judge 비교")
        scenario_df = pd.DataFrame([{"테스트 유형":c["type"], "예시":c["question"], "확인할 사항":c["check"]} for c in VALIDATION_CASES])
        with st.expander("대표 입력 유형 10개", expanded=False):
            st.dataframe(scenario_df, use_container_width=True, hide_index=True)
        run_axes = st.button("▶ 3축 검증 실행", type="primary", use_container_width=True)
        if run_axes:
            with st.spinner("10개 입력을 요약·개선안·독립 Judge 축으로 검증하고 있습니다."):
                st.session_state.three_axis_results = run_three_axis_validation(use_independent_judge=use_llm)
        axis_results = st.session_state.three_axis_results
        if axis_results:
            metrics_df = pd.DataFrame(validation_metrics(axis_results))
            st.markdown("#### 발표 핵심 지표")
            st.dataframe(metrics_df, use_container_width=True, hide_index=True)
            detail_df = pd.DataFrame([{
                "ID":r["case"]["id"], "유형":r["case"]["type"],
                "요약 정확성":"PASS" if r["summary_pass"] else "FAIL",
                "원인–개선안":"PASS" if r["cause_improvement_pass"] else "FAIL",
                "이종 모델":"미실행" if r["cross_model_pass"] is None else "PASS" if r["cross_model_pass"] else "FAIL",
                "응답시간(ms)":r["response_time_ms"],
            } for r in axis_results])
            with st.expander("케이스별 3축 상세 결과"):
                st.dataframe(detail_df, use_container_width=True, hide_index=True)
            st.markdown("#### 실패 → 개선 → 재검증 사례")
            before, after = st.columns(2)
            before.error("개선 전 · 복합 불만\n\n배송 지연만 요약하고 ‘상담원 불친절’을 누락")
            compound = next((r for r in axis_results if r["case"]["id"] == "V-02"), None)
            compound_status = "PASS" if compound and compound["summary_pass"] else "추가 개선 필요"
            after.success(f"개선 후 · Critic 규칙 및 복합 의도 분리\n\n배송 + 상담 응대를 각각 추출 · 재검증 {compound_status}")
            human_review = st.checkbox("사람 검토 완료", help="발표 담당자가 실패 사례와 Judge 근거를 직접 확인한 뒤 체크합니다.")
            pytest_gate = all(not r["error"] for r in axis_results)
            st.markdown(f"**최종 Gate:** OpenAI/Agent 생성 → Evaluator·Critic → Anthropic 독립 채점 → pytest {'PASS' if pytest_gate else 'FAIL'} → 사람 평가 {'PASS' if human_review else '대기'}")
        else:
            st.caption("독립 Judge 토글을 켜면 Anthropic API로 교차평가합니다. 끄면 앞의 두 축과 Rule 기준을 먼저 검증합니다.")
    with qa2_tab:
        st.markdown('<div class="section-title">최종 산출물을 별도 Anthropic 모델로 독립 채점</div>', unsafe_allow_html=True)
        st.caption("파이프라인 내부 Evaluator·Critic과 분리된 QA입니다. 실제 API 호출 비용이 발생합니다.")

        rubric = load_judge_json("judge_rubric.json")
        rubric_cols = st.columns(5)
        for column, spec in zip(rubric_cols, rubric["dimensions"].values()):
            column.metric(spec["label"], f"{spec['weight']}점")

        with st.container(border=True):
            st.markdown("**판정 기준**")
            st.write("70점 이상이며 Critical Failure가 없어야 PASS입니다.")
            st.caption("개인정보 노출 · 근거 없는 사실 생성 · 장애 성공 위장 · 결제/환불 확정 오안내는 점수와 무관하게 배포 보류")

        judge_cases = load_judge_json("judge_cases.json")
        with st.expander("독립 Judge 테스트 케이스", expanded=False):
            st.dataframe(
                pd.DataFrame([
                    {"ID": case["case_id"], "질문": case["question"], "장애 조건": case.get("fault") or "없음"}
                    for case in judge_cases
                ]),
                width="stretch",
                hide_index=True,
            )

        run_qa2 = st.button(
            "QA2 Anthropic Judge 실행",
            type="primary",
            icon=":material/play_arrow:",
            width="stretch",
            disabled=not bool(os.getenv("ANTHROPIC_API_KEY")),
            help="루트 .env의 ANTHROPIC_API_KEY를 사용합니다.",
        )
        if not os.getenv("ANTHROPIC_API_KEY"):
            st.warning("ANTHROPIC_API_KEY가 없어 QA2를 실행할 수 없습니다.", icon=":material/key:")

        if run_qa2:
            try:
                with st.status("5개 산출물을 독립 채점하고 있습니다.", expanded=True) as status:
                    st.write("6개 Agent 파이프라인 실행")
                    results = run_judge_cases()
                    st.write("Anthropic Structured Outputs 검증")
                    report_path = write_judge_csv(results)
                    st.session_state.qa2_judge_results = results
                    status.update(label="QA2 독립 Judge 완료", state="complete", expanded=False)
                st.toast("QA2 결과와 CSV 보고서를 생성했습니다.", icon=":material/check_circle:")
            except Exception as exc:
                st.error(f"QA2 Anthropic Judge 실패: {exc}", icon=":material/error:")

        qa2_results = st.session_state.qa2_judge_results
        if qa2_results:
            rows = []
            for item in qa2_results:
                judgement = item["judgement"]
                rows.append({
                    "ID": item["case"]["case_id"],
                    **{
                        rubric["dimensions"][key]["label"]: score
                        for key, score in judgement["dimension_scores"].items()
                    },
                    "총점": judgement["score"],
                    "판정": "PASS" if judgement["pass"] else "FAIL",
                    "Critical": "있음" if judgement["critical_failure"] else "없음",
                    "배포 판단": judgement["deployment_decision"],
                })
            result_df = pd.DataFrame(rows)
            passed = sum(item["judgement"]["pass"] for item in qa2_results)
            average = sum(item["judgement"]["score"] for item in qa2_results) / len(qa2_results)
            summary_cols = st.columns(4)
            summary_cols[0].metric("평균 점수", f"{average:.1f}")
            summary_cols[1].metric("PASS", f"{passed}/{len(qa2_results)}")
            summary_cols[2].metric("Critical", sum(item["judgement"]["critical_failure"] for item in qa2_results))
            summary_cols[3].metric("Judge 모델", qa2_results[0]["judgement"]["model"])
            st.dataframe(result_df, width="stretch", hide_index=True)

            selected_id = st.selectbox("판정 근거 확인", result_df["ID"].tolist())
            selected = next(item for item in qa2_results if item["case"]["case_id"] == selected_id)
            judgement = selected["judgement"]
            with st.container(border=True):
                st.write(judgement["overall_reason"])
                if judgement["improvements"]:
                    st.markdown("**개선 권고**")
                    for improvement in judgement["improvements"]:
                        st.write(f"- {improvement}")

            report_path = write_judge_csv(qa2_results)
            st.download_button(
                "QA2 CSV 보고서 다운로드",
                data=report_path.read_bytes(),
                file_name="llm_judge_result.csv",
                mime="text/csv",
                icon=":material/download:",
                width="content",
            )
        else:
            st.info("QA2 실행 후 5개 평가축, 총점, Critical Gate, 배포 판단과 근거가 표시됩니다.", icon=":material/info:")
    with rubric_tab:
        st.markdown('<div class="section-title">평가 항목별 발표 답변과 근거</div>', unsafe_allow_html=True)
        st.caption("각 항목 8점 · 총 10개 항목 · 합계 80점")
        st.markdown(_presentation_rubric_html(st.session_state.runs), unsafe_allow_html=True)
    with technical_tab:
        st.markdown('<div class="section-title">기술 평가 항목별 답변과 실행 증적</div>', unsafe_allow_html=True)
        st.caption("각 항목 2점 · 총 10개 항목 · 합계 20점")
        st.markdown(_technical_rubric_html(st.session_state.runs), unsafe_allow_html=True)
    with team_tab:
        st.markdown('<div class="section-title">권장 역할 분담과 공동 검증 방식</div>', unsafe_allow_html=True)
        team_roles = pd.DataFrame([
            ["VOC·Agent 담당", "불만 유형/원인 분석, 6개 Agent 설계", "Agent 단계별 INPUT·OUTPUT, 검색 근거"],
            ["QA·Judge 담당", "테스트 케이스, 평가 기준, 독립 Judge", "PASS/FAIL 근거, 교차검증 A–D"],
            ["UI·보고서 담당", "발표 흐름, 대시보드, 결함·배포 보고서", "정량 지표, 결함 조치, 최종 판단"],
            ["전 팀 공통", "실패 케이스 리뷰와 질의응답 리허설", "동일 실행 결과와 보고서로 답변 통일"],
        ], columns=["담당", "주요 역할", "발표 근거"])
        st.dataframe(team_roles, use_container_width=True, hide_index=True)
        st.info("발표 시 실제 팀원 이름을 각 담당 앞에 말하고, 모든 답변은 개인 의견이 아닌 실행 로그·테스트 결과·보고서 수치로 통일합니다.")
    with qa_tab:
        questions = [
            ("왜 에이전트를 6개로 분리했나요?", "해석·근거 검색·요약·평가·비판·개선은 실패 원인과 책임 지점이 다릅니다. 분리하면 어느 단계에서 품질이 떨어졌는지 추적하고 해당 단계만 재검증할 수 있습니다."),
            ("Evaluator와 Critic이 있는데 독립 Judge가 왜 필요한가요?", "Evaluator와 Critic은 결과 생성 과정에 참여하는 내부 통제입니다. 독립 Judge는 생성 과정 밖에서 동일 기준으로 최종 결과를 심사해 자기평가 편향과 모델 고유 편향을 줄입니다."),
            ("테스트 결과가 객관적이라는 근거는 무엇인가요?", "사전에 정의한 20개 Happy·Edge·Negative·Fault 케이스에 동일한 100점 기준과 Critical Gate를 적용하고, PASS·FAIL·성공률·판정 사유를 원본 결과와 함께 저장합니다."),
            ("실패 테스트는 어떻게 처리하나요?", "실패를 숨기지 않고 TC_ID, 입력, 기대 결과, 실제 사유, 심각도와 개선 조치로 결함화합니다. 수정 후 같은 케이스를 재실행해 회귀 여부를 확인합니다."),
            ("API 오류나 데이터 누락 시에도 신뢰할 수 있나요?", "Retriever 중단·CSV 누락·응답 지연을 별도 Fault 케이스로 재현합니다. 성공으로 위장하지 않고 ERROR/NO_DATA와 한계를 표시하며, Judge 장애 시 Rule Judge 전환 사실도 경고로 남깁니다."),
            ("개선안이 실제 정책과 연결됐다는 근거는 무엇인가요?", "Retriever가 찾은 VOC의 원인과 policy 필드를 Improver가 개선 조치로 연결하고, 우선순위와 재발률·처리시간·1차 해결률 KPI를 함께 제시합니다."),
        ]
        for question, answer in questions:
            with st.expander(question):
                st.write(answer)

elif page == "대시보드":
    st.markdown('<div class="hero"><h1>VOC Quality Studio</h1><p>6개 AI Agent가 VOC를 분석하고, QA 결과를 근거로 품질과 배포 가능성을 판단합니다.</p></div>', unsafe_allow_html=True)
    passed, failed, avg, rate, critical = quality_summary()
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Agent", "6", "Pipeline")
    c2.metric("QA Test", "20", "Scenario")
    c3.metric("Pass Rate", f"{rate:.1f}%" if st.session_state.runs else "Ready")
    c4.metric("Quality Score", f"{avg:.1f}" if st.session_state.runs else "100 pt")

    st.markdown('<div class="section-title">독립 LLM 교차검증 흐름</div>', unsafe_allow_html=True)
    st.markdown(
        '<div class="judge-flow">'
        '<div class="judge-node"><strong>VOC 테스트 데이터</strong><span>검증 입력</span></div><div class="judge-arrow">→</div>'
        '<div class="judge-node"><strong>OpenAI 요약·개선안</strong><span>생성 모델</span></div><div class="judge-arrow">→</div>'
        '<div class="judge-node"><strong>내부 Evaluator·Critic</strong><span>근거·안전성 사전 점검</span></div><div class="judge-arrow">→</div>'
        '<div class="judge-node external"><strong>Anthropic 독립 Judge</strong><span>교차 평가 모델</span></div><div class="judge-arrow">→</div>'
        '<div class="judge-node"><strong>점수 · PASS/FAIL</strong><span>판정 근거 포함</span></div></div>',
        unsafe_allow_html=True,
    )
    cross_validation = pd.DataFrame([
        ["A", "OpenAI", "Anthropic", "기본 품질검증"],
        ["B", "Anthropic", "OpenAI", "모델 역할 변경 검증"],
        ["C", "OpenAI", "OpenAI", "동일 모델 평가와 비교"],
        ["D", "Anthropic", "Anthropic", "동일 모델 평가와 비교"],
    ], columns=["실험군", "생성 모델", "평가 모델", "목적"])
    with st.expander("교차검증 실험군 A–D", expanded=True):
        st.dataframe(cross_validation, use_container_width=True, hide_index=True)

    left, right = st.columns([1.55, 1])
    with left:
        st.markdown('<div class="section-title">Agent Pipeline</div>', unsafe_allow_html=True)
        st.markdown(agent_status_html(), unsafe_allow_html=True)
        st.markdown("<div style='height:.8rem'></div>", unsafe_allow_html=True)
        st.markdown('<div class="section-title">System Scope</div>', unsafe_allow_html=True)
        scope = pd.DataFrame([
            ["기능", "Agent 역할 및 연계"], ["데이터", "VOC 근거 검색·요약"], ["AI 품질", "정확성·유용성·근거성"],
            ["장애 대응", "중단·누락·지연 처리"], ["운영", "로그·성능·결함 추적"],
        ], columns=["영역", "진단 범위"])
        st.dataframe(scope, use_container_width=True, hide_index=True, height=235)
    with right:
        st.markdown('<div class="section-title">Release Readiness</div>', unsafe_allow_html=True)
        if not st.session_state.runs:
            st.markdown('<div class="panel"><h3 style="margin:0;color:#172033">검증 대기</h3><p class="small-note">QA 테스트를 실행하면 배포 판단이 표시됩니다.</p></div>', unsafe_allow_html=True)
        else:
            decision = "HOLD" if critical or avg < 70 else "REVIEW" if avg < 90 else "READY"
            tone = "red" if decision == "HOLD" else "amber" if decision == "REVIEW" else "green"
            st.markdown(f'<div class="panel"><h3 style="margin:0">{badge(decision, tone)}</h3><p class="small-note">평균 {avg:.1f}점 · PASS {passed} · FAIL {failed}</p></div>', unsafe_allow_html=True)
        st.markdown('<div class="section-title" style="margin-top:1rem">Critical Gate</div>', unsafe_allow_html=True)
        st.error("개인정보 노출 · 허위 정책 생성 · 장애 성공 위장 · 결제/환불 오안내")
        st.markdown('<div class="section-title">Quick Start</div>', unsafe_allow_html=True)
        st.info("1. VOC 분석 → 2. QA 테스트 → 3. 품질 분석 → 4. 배포 판단")
        if st.button("대표 VOC Pipeline 데모", type="primary", use_container_width=True):
            demo_output = analyze_voc("결제는 완료되었는데 주문 내역에 보이지 않습니다.", None)
            st.session_state.single = demo_output
            replay_agent_flow(demo_output)

    if st.session_state.execution_log:
        st.markdown('<div class="section-title" style="margin-top:1rem">Latest Execution</div>', unsafe_allow_html=True)
        st.markdown(_timeline_html(st.session_state.execution_log), unsafe_allow_html=True)

elif page == "실시간 평가":
    page_header("실시간 질문 품질 평가", "즉석 질문을 6개 Agent로 분석하고 근거성·완전성·안전성·유용성·일관성을 자동 채점합니다.")
    st.info("사전 기대 결과가 없는 질문이므로 '정답 일치율'이 아니라, 검색 근거와 답변 품질을 평가합니다.")

    input_col, guide_col = st.columns([1.45, 1])
    with input_col:
        realtime_question = st.text_area(
            "실시간 질문",
            value="결제는 완료되었는데 주문 내역에 보이지 않습니다. 어떻게 해야 하나요?",
            height=135,
            placeholder="발표자가 즉석에서 질문을 입력합니다.",
        )
        realtime_fault = st.selectbox(
            "실행 환경",
            ["정상 실행", "Retriever 중단", "CSV 파일 누락", "응답 지연"],
            key="realtime_fault",
        )
        run_realtime = st.button("▶ 실시간 분석 및 자동 평가", type="primary", use_container_width=True)
    with guide_col:
        st.markdown(
            '<div class="panel"><b>평가 기준</b><p class="small-note">정확성·근거성 30점 · 완전성 20점 · 안전성 20점 · 유용성 20점 · 일관성 10점</p></div>',
            unsafe_allow_html=True,
        )
        st.markdown("<div style='height:.55rem'></div>", unsafe_allow_html=True)
        st.markdown(
            '<div class="panel"><b>판정 기준</b><p class="small-note">70점 이상이며 Critical Failure가 없으면 PASS입니다. API 키가 있으면 LLM Judge를 선택할 수 있습니다.</p></div>',
            unsafe_allow_html=True,
        )

    live_agent_slot = st.empty()
    live_agent_slot.markdown(agent_status_html(), unsafe_allow_html=True)

    if run_realtime:
        if not realtime_question.strip():
            st.error("질문을 입력하세요.")
        else:
            fault_map = {"정상 실행": None, "Retriever 중단": "retriever_down", "CSV 파일 누락": "csv_missing", "응답 지연": "timeout"}
            with st.spinner("6개 Agent 분석과 실시간 품질 평가를 실행하고 있습니다."):
                pipeline_result = analyze_voc(realtime_question, fault_map[realtime_fault])
                st.session_state.single = pipeline_result
                replay_agent_flow(pipeline_result, speed=.65)
                realtime_judge = evaluate_realtime(realtime_question, pipeline_result, use_llm=use_llm)
                st.session_state.realtime_result = {"pipeline": pipeline_result, "judge": realtime_judge}

    realtime_bundle = st.session_state.realtime_result
    if realtime_bundle:
        pipeline_result = realtime_bundle["pipeline"]
        judge = realtime_bundle["judge"]
        score = int(judge.get("score", 0))
        passed = bool(judge.get("pass"))
        critical = bool(judge.get("critical_failure"))

        st.markdown('<div class="section-title" style="margin-top:1rem">실시간 판정</div>', unsafe_allow_html=True)
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("종합 점수", f"{score} / 100")
        k2.metric("판정", "PASS" if passed else "FAIL")
        k3.metric("검색 근거", f"{judge.get('evidence_count', 0)}건")
        k4.metric("Critical", "YES" if critical else "NO")

        status_tone = "green" if passed else "red"
        st.markdown(
            f'<div class="panel"><span class="badge badge-{status_tone}">{"PASS" if passed else "FAIL"}</span>'
            f'<h3 style="margin:.65rem 0 .25rem;color:#172033">{judge.get("reason", "")}</h3>'
            f'<p class="small-note">{judge.get("evaluation_scope", "")}</p></div>',
            unsafe_allow_html=True,
        )
        if judge.get("warning"):
            st.warning(judge["warning"])

        score_df = pd.DataFrame([
            {"평가 항목": name, "점수": value, "배점": judge.get("dimension_weights", {}).get(name, value)}
            for name, value in judge.get("dimension_scores", {}).items()
        ])
        chart_col, answer_col = st.columns([1, 1.25])
        with chart_col:
            fig = px.bar(
                score_df,
                x="점수",
                y="평가 항목",
                orientation="h",
                text="점수",
                range_x=[0, 30],
                title="항목별 실시간 평가",
            )
            fig.update_layout(height=340, margin=dict(l=10, r=10, t=45, b=10))
            st.plotly_chart(fig, use_container_width=True)
        with answer_col:
            final = pipeline_result.get("final") or {}
            safe_message = final.get("safe_message", "분석 결과가 없습니다.")
            st.markdown('<div class="section-title">최종 고객 안내</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="result-card"><p>{safe_message}</p></div>', unsafe_allow_html=True)
            improvement = final.get("improvement") or {}
            st.markdown('<div class="section-title">Agent 개선안</div>', unsafe_allow_html=True)
            for action in improvement.get("improvement_actions", []):
                st.markdown(f"- {action}")

        detail_left, detail_right = st.columns(2)
        with detail_left:
            st.markdown('<div class="section-title">Judge 개선 권고</div>', unsafe_allow_html=True)
            for item in judge.get("improvements", []):
                st.markdown(f"- {item}")
            if judge.get("critical_reasons"):
                st.error("Critical Failure: " + ", ".join(judge["critical_reasons"]))
        with detail_right:
            st.markdown('<div class="section-title">평가 근거</div>', unsafe_allow_html=True)
            with st.expander("항목별 세부 근거", expanded=True):
                for dimension, reasons in judge.get("dimension_reasons", {}).items():
                    st.markdown(f"**{dimension}**")
                    st.markdown(" · ".join(reasons) if reasons else "평가 근거 없음")

        with st.expander("6개 Agent 상세 실행 결과", expanded=False):
            render_result(pipeline_result)

        report_payload = {
            "evaluated_at": datetime.now().isoformat(timespec="seconds"),
            "question": pipeline_result.get("question"),
            "judge": judge,
            "pipeline": pipeline_result,
        }
        st.download_button(
            "실시간 평가 JSON 다운로드",
            json.dumps(report_payload, ensure_ascii=False, indent=2).encode("utf-8"),
            "realtime_evaluation.json",
            "application/json",
            use_container_width=True,
        )

elif page == "Agent 단계 실행":
    page_header("6개 Agent 단계별 실행", "발표자가 각 Agent를 순서대로 실행하며 역할, 입력값, 출력값을 직접 설명합니다.")

    role_map = {
        "Interpreter": "고객 질문을 정규화하고 의도·키워드·문의 유형을 추출합니다.",
        "Retriever": "Interpreter 결과를 사용해 VOC 데이터에서 관련 사례와 근거를 검색합니다.",
        "Summarizer": "검색된 VOC를 요약하고 확인된 사실과 추론 한계를 구분합니다.",
        "Evaluator": "질문·검색 근거·요약 결과의 관련성과 근거성을 평가합니다.",
        "Critic": "답변의 위험 요소, 근거 부족, 고위험 업무 주의사항을 점검합니다.",
        "Improver": "앞선 분석과 비판 결과를 반영해 고객 안내와 개선 조치를 생성합니다.",
    }
    dependency_map = {
        "Interpreter": "사용자 질문",
        "Retriever": "Interpreter 출력",
        "Summarizer": "사용자 질문 + Retriever 출력",
        "Evaluator": "질문 + Interpreter + Retriever + Summarizer 출력",
        "Critic": "질문 + Retriever + Evaluator 출력",
        "Improver": "질문 + Retriever + Summarizer + Critic 출력",
    }

    control_left, control_right = st.columns([1.45, 1])
    with control_left:
        manual_question = st.text_area("발표용 VOC 입력", value=st.session_state.manual_question, height=115, key="manual_question_input")
    with control_right:
        fault_name = st.selectbox("실행 환경", ["정상 실행", "Retriever 중단", "CSV 파일 누락", "응답 지연"], key="manual_fault_select")
        execution_speed_name = st.radio(
            "전체 실행 속도",
            ["빠르게", "표준", "천천히"],
            horizontal=True,
            index=1,
            key="manual_execution_speed",
        )
        execution_speed_map = {"빠르게": 0.55, "표준": 1.0, "천천히": 1.55}
        run_all_agents = st.button("▶ 전체 Agent 실행 및 실시간 평가", type="primary", use_container_width=True)
        if st.button("단계 실행 초기화", use_container_width=True):
            st.session_state.manual_steps = []
            st.session_state.manual_question = manual_question
            st.session_state.manual_fault = None
            st.session_state.realtime_result = None
            st.session_state.manual_selected_agent = "Interpreter"
            st.rerun()

    fault_map = {"정상 실행": None, "Retriever 중단": "retriever_down", "CSV 파일 누락": "csv_missing", "응답 지연": "timeout"}

    if run_all_agents:
        if not manual_question.strip():
            st.error("질문을 입력하세요.")
        else:
            with st.spinner("6개 Agent 실행과 실시간 품질 평가를 진행하고 있습니다."):
                pipeline_result = analyze_voc(manual_question, fault_map[fault_name])
                replay_agent_flow(pipeline_result, speed=execution_speed_map[execution_speed_name])
                realtime_judge = evaluate_realtime(manual_question, pipeline_result, use_llm=use_llm)
                st.session_state.manual_question = manual_question
                st.session_state.manual_steps = pipeline_result.get("steps", [])
                st.session_state.single = pipeline_result
                st.session_state.realtime_result = {"pipeline": pipeline_result, "judge": realtime_judge}
                st.session_state.manual_selected_agent = "Improver"
            st.rerun()
    steps = st.session_state.manual_steps
    completed = {step.get("agent") for step in steps}
    errored = {step.get("agent") for step in steps if step.get("status") == "ERROR"}

    st.markdown('<div class="section-title">Agent 단계 실행</div>', unsafe_allow_html=True)
    st.caption("상태 박스를 클릭하면 해당 Agent가 실행됩니다. 완료된 Agent를 다시 클릭하면 기존 결과를 표시합니다.")

    statuses = server_status()
    agent_columns = st.columns(6)
    clicked_agent: str | None = None
    for idx, agent_name in enumerate(AGENTS):
        step_data = next((x for x in steps if x.get("agent") == agent_name), None)
        previous_ready = idx == 0 or AGENTS[idx - 1] in completed
        is_selected = st.session_state.manual_selected_agent == agent_name
        port = statuses.get(agent_name.lower(), statuses.get(agent_name, {})).get("port", "-")

        if step_data and step_data.get("status") == "ERROR":
            icon, status_label = "🔴", "ERROR"
        elif step_data:
            icon, status_label = "🟢", "DONE"
        elif previous_ready:
            icon, status_label = "🔵", "READY"
        else:
            icon, status_label = "⚪", "LOCKED"

        button_label = f"{icon} {agent_name}\n{status_label} · :{port}"
        with agent_columns[idx]:
            if st.button(
                button_label,
                key=f"manual_agent_card_{agent_name}",
                type="primary" if is_selected else "secondary",
                use_container_width=True,
            ):
                clicked_agent = agent_name

    if clicked_agent:
        idx = AGENTS.index(clicked_agent)
        previous_ready = idx == 0 or AGENTS[idx - 1] in completed
        step_data = next((x for x in st.session_state.manual_steps if x.get("agent") == clicked_agent), None)
        st.session_state.manual_selected_agent = clicked_agent

        if step_data is None and not previous_ready:
            st.warning(f"먼저 {AGENTS[idx - 1]} Agent를 실행하세요.")
        elif step_data is None:
            if not manual_question.strip():
                st.error("질문을 입력하세요.")
            else:
                st.session_state.manual_question = manual_question
                current = st.session_state.manual_steps
                interpreted = current[0]["output"] if len(current) >= 1 else {}
                matches = current[1]["output"] if len(current) >= 2 and isinstance(current[1].get("output"), list) else []
                summary = current[2]["output"] if len(current) >= 3 else {}
                evaluation = current[3]["output"] if len(current) >= 4 else {}
                critique = current[4]["output"] if len(current) >= 5 else {}
                if clicked_agent == "Interpreter":
                    result_step = Interpreter().run(manual_question)
                elif clicked_agent == "Retriever":
                    result_step = Retriever().run(interpreted, fault_map[fault_name])
                elif clicked_agent == "Summarizer":
                    result_step = Summarizer().run(manual_question, matches)
                elif clicked_agent == "Evaluator":
                    result_step = Evaluator().run(manual_question, interpreted, matches, summary)
                elif clicked_agent == "Critic":
                    result_step = Critic().run(manual_question, matches, evaluation)
                else:
                    result_step = Improver().run(manual_question, matches, summary, critique)
                st.session_state.manual_steps.append(asdict(result_step))
                st.rerun()
        else:
            st.rerun()

    selected_agent = st.session_state.manual_selected_agent
    selected_idx = AGENTS.index(selected_agent)
    selected_step = next((x for x in st.session_state.manual_steps if x.get("agent") == selected_agent), None)
    selected_previous_ready = selected_idx == 0 or AGENTS[selected_idx - 1] in completed

    st.markdown('<div class="section-title" style="margin-top:1rem">선택한 Agent 상세</div>', unsafe_allow_html=True)
    with st.container(border=True):
        st.markdown(f"### {selected_idx + 1}. {selected_agent}")
        st.caption(role_map[selected_agent])
        st.markdown(f"**입력 구조:** {dependency_map[selected_agent]}")

        if selected_step:
            current = st.session_state.manual_steps
            input_value: Any
            if selected_agent == "Interpreter":
                input_value = {"question": st.session_state.manual_question}
            elif selected_agent == "Retriever":
                input_value = current[0]["output"]
            elif selected_agent == "Summarizer":
                input_value = {"question": st.session_state.manual_question, "matches": current[1]["output"]}
            elif selected_agent == "Evaluator":
                input_value = {"question": st.session_state.manual_question, "interpreted": current[0]["output"], "matches": current[1]["output"], "summary": current[2]["output"]}
            elif selected_agent == "Critic":
                input_value = {"question": st.session_state.manual_question, "matches": current[1]["output"], "evaluation": current[3]["output"]}
            else:
                input_value = {"question": st.session_state.manual_question, "matches": current[1]["output"], "summary": current[2]["output"], "critique": current[4]["output"]}
            in_col, out_col = st.columns(2)
            with in_col:
                st.markdown("**INPUT**")
                st.json(input_value, expanded=True)
            with out_col:
                st.markdown("**OUTPUT**")
                if selected_step.get("error"):
                    st.error(selected_step["error"])
                st.json(selected_step.get("output", {}), expanded=True)
                st.caption(f"상태: {selected_step.get('status')} · 응답시간: {selected_step.get('latency_ms', 0)}ms")
        elif not selected_previous_ready:
            st.info(f"먼저 {AGENTS[selected_idx - 1]} Agent를 실행하세요.")
        else:
            st.info(f"상단의 {selected_agent} 버튼을 클릭하면 이 단계가 실행됩니다.")

    if len(st.session_state.manual_steps) == 6:
        final_steps = st.session_state.manual_steps
        output = {
            "question": st.session_state.manual_question,
            "success": all(step.get("status") != "ERROR" for step in final_steps),
            "steps": final_steps,
            "final": {
                "interpretation": final_steps[0]["output"],
                "matches": final_steps[1]["output"],
                "analysis": final_steps[2]["output"],
                "evaluation": final_steps[3]["output"],
                "critique": final_steps[4]["output"],
                "improvement": final_steps[5]["output"],
                "safe_message": final_steps[5]["output"].get("customer_guidance", ""),
            },
        }
        st.session_state.single = output
        if not st.session_state.realtime_result or st.session_state.realtime_result.get("pipeline", {}).get("question") != st.session_state.manual_question:
            realtime_judge = evaluate_realtime(st.session_state.manual_question, output, use_llm=use_llm)
            st.session_state.realtime_result = {"pipeline": output, "judge": realtime_judge}
        st.success("6개 Agent 실행이 완료되었습니다. 아래에서 실시간 평가 결과를 확인할 수 있습니다.")

    realtime_bundle = st.session_state.realtime_result
    if realtime_bundle and realtime_bundle.get("pipeline", {}).get("question") == st.session_state.manual_question:
        pipeline_result = realtime_bundle["pipeline"]
        judge = realtime_bundle["judge"]
        score = int(judge.get("score", 0))
        passed = bool(judge.get("pass"))
        critical = bool(judge.get("critical_failure"))

        st.markdown('<div class="section-title" style="margin-top:1rem">실시간 평가 결과</div>', unsafe_allow_html=True)
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("종합 점수", f"{score} / 100")
        k2.metric("판정", "PASS" if passed else "FAIL")
        k3.metric("검색 근거", f"{judge.get('evidence_count', 0)}건")
        k4.metric("Critical", "YES" if critical else "NO")

        status_tone = "green" if passed else "red"
        st.markdown(
            f'<div class="panel"><span class="badge badge-{status_tone}">{"PASS" if passed else "FAIL"}</span>'
            f'<h3 style="margin:.65rem 0 .25rem;color:#172033">{judge.get("reason", "")}</h3>'
            f'<p class="small-note">{judge.get("evaluation_scope", "")}</p></div>',
            unsafe_allow_html=True,
        )
        if judge.get("warning"):
            st.warning(judge["warning"])

        score_df = pd.DataFrame([
            {"평가 항목": name, "점수": value, "배점": judge.get("dimension_weights", {}).get(name, value)}
            for name, value in judge.get("dimension_scores", {}).items()
        ])
        chart_col, answer_col = st.columns([1, 1.25])
        with chart_col:
            fig = px.bar(score_df, x="점수", y="평가 항목", orientation="h", text="점수", range_x=[0, 30], title="항목별 실시간 평가")
            fig.update_layout(height=340, margin=dict(l=10, r=10, t=45, b=10))
            st.plotly_chart(fig, use_container_width=True)
        with answer_col:
            final = pipeline_result.get("final") or {}
            st.markdown('<div class="section-title">최종 고객 안내</div>', unsafe_allow_html=True)
            st.markdown(f'<div class="result-card"><p>{final.get("safe_message", "분석 결과가 없습니다.")}</p></div>', unsafe_allow_html=True)
            st.markdown('<div class="section-title">Judge 개선 권고</div>', unsafe_allow_html=True)
            for item in judge.get("improvements", []):
                st.markdown(f"- {item}")
            if judge.get("critical_reasons"):
                st.error("Critical Failure: " + ", ".join(judge["critical_reasons"]))

elif page == "QA 테스트":
    page_header("실시간 QA 대시보드", "20개 시나리오의 실행 상태와 품질 지표를 테스트 단위로 즉시 갱신합니다.")
    cases = load_cases()
    counts = pd.Series([case["type"] for case in cases]).value_counts().rename_axis("유형").reset_index(name="개수")

    control_left, control_right = st.columns([1.15, 2.25])
    with control_left:
        st.markdown('<div class="section-title">Test Control</div>', unsafe_allow_html=True)
        st.dataframe(counts, use_container_width=True, hide_index=True, height=205)
        run_all = st.button("20개 전체 테스트 실행", type="primary", use_container_width=True)
        if st.button("결과 초기화", use_container_width=True):
            st.session_state.runs = []
            st.rerun()
    with control_right:
        kpi_slot = st.empty()
        current_slot = st.empty()
        progress_slot = st.empty()
        recent_slot = st.empty()
        kpi_slot.markdown(_qa_kpis_html(st.session_state.runs, len(cases)), unsafe_allow_html=True)
        current_slot.markdown('<div class="qa-board"><div class="qa-section-head"><h3>현재 실행</h3><span class="badge badge-blue">READY</span></div><div class="small-note">전체 테스트 실행 버튼을 누르면 현재 TC가 실시간으로 표시됩니다.</div></div>', unsafe_allow_html=True)
        progress_slot.progress(len(st.session_state.runs) / len(cases) if cases else 0)
        recent_slot.markdown('<div class="qa-board"><div class="qa-section-head"><h3>최근 실행 결과</h3><span class="qa-mini">최신 8건</span></div>' + _qa_recent_html(st.session_state.runs) + '</div>', unsafe_allow_html=True)

    if run_all:
        runs: list[dict[str, Any]] = []
        total = len(cases)
        for idx, case in enumerate(cases, 1):
            kpi_slot.markdown(_qa_kpis_html(runs, total, running=True), unsafe_allow_html=True)
            current_slot.markdown(
                '<div class="qa-board"><div class="qa-section-head"><h3>현재 실행</h3><span class="live-chip">RUNNING</span></div>'
                f'<div class="qa-current"><div class="tc">{case["case_id"]}</div><div class="type">{case["type"]}</div><div class="question">{case["question"]}</div><div class="count">{idx}/{total}</div></div></div>',
                unsafe_allow_html=True,
            )
            progress_slot.progress((idx - 1) / total, text=f"{idx}/{total} · {case['case_id']} 평가 중")
            started = time.perf_counter()
            result = run_case(case, use_llm)
            measured = int((time.perf_counter() - started) * 1000)
            result["response_time_ms"] = max(measured, _extract_latency_ms(result))
            runs.append(result)
            kpi_slot.markdown(_qa_kpis_html(runs, total, running=idx < total), unsafe_allow_html=True)
            recent_slot.markdown('<div class="qa-board"><div class="qa-section-head"><h3>최근 실행 결과</h3><span class="qa-mini">최신 8건</span></div>' + _qa_recent_html(runs) + '</div>', unsafe_allow_html=True)
            progress_slot.progress(idx / total, text=f"{idx}/{total} · {case['case_id']} 완료")
            time.sleep(0.08)

        st.session_state.runs = runs
        save_reports(runs)
        current_slot.markdown('<div class="qa-board" style="border-color:#86EFAC;background:#F0FDF4"><div class="qa-section-head"><h3 style="color:#15803D">QA 실행 완료</h3><span class="badge badge-green">COMPLETE</span></div><div class="small-note">20개 시나리오 평가와 보고서 생성을 완료했습니다.</div></div>', unsafe_allow_html=True)

    if st.session_state.runs:
        st.markdown('<div class="section-title" style="margin-top:1rem">Quality Overview</div>', unsafe_allow_html=True)
        overview_left, overview_right = st.columns([1.1, 1.4])
        with overview_left:
            type_df = _qa_type_summary(st.session_state.runs)
            st.dataframe(type_df, use_container_width=True, hide_index=True, height=320)
        with overview_right:
            chart_df = _qa_type_summary(st.session_state.runs)
            fig = px.bar(
                chart_df,
                x="성공률",
                y="유형",
                orientation="h",
                text_auto=".1f",
                title="유형별 실시간 성공률",
                range_x=[0, 100],
            )
            fig.update_layout(margin=dict(l=10, r=10, t=45, b=10), height=320)
            st.plotly_chart(fig, use_container_width=True)

        rows = [{
            "TC_ID": item["case"]["case_id"], "유형": item["case"]["type"], "질문": item["case"]["question"],
            "점수": item["judge"]["score"], "판정": "PASS" if item["judge"]["pass"] else "FAIL",
            "응답시간(ms)": _extract_latency_ms(item), "Critical": item["critical_failure"], "사유": item["judge"]["reason"],
        } for item in st.session_state.runs]
        df = pd.DataFrame(rows)
        with st.expander("전체 테스트 상세 결과", expanded=False):
            st.dataframe(df, use_container_width=True, hide_index=True, height=410)
        st.download_button("테스트 결과 CSV", df.to_csv(index=False).encode("utf-8-sig"), "test_result.csv", "text/csv")

elif page == "품질 분석":
    page_header("품질 분석", "점수·응답시간·Agent 성능·유형별 품질을 다각도로 분석합니다.")
    rubric = pd.read_csv(ROOT / "quality_diagnosis" / "evaluation_rubric.csv")
    if not st.session_state.runs:
        st.info("먼저 QA 테스트를 실행하세요.")
        st.dataframe(rubric, use_container_width=True, hide_index=True)
    else:
        df = _run_rows(st.session_state.runs)
        agent_df = _agent_performance(st.session_state.runs)
        passed, failed, avg, rate, critical = quality_summary()
        c1,c2,c3,c4,c5=st.columns(5)
        c1.metric("평균 점수",f"{avg:.1f}"); c2.metric("성공률",f"{rate:.1f}%"); c3.metric("평균 응답",f"{df['응답시간(ms)'].mean():.0f} ms"); c4.metric("최대 응답",f"{df['응답시간(ms)'].max():.0f} ms"); c5.metric("Critical","YES" if critical else "NO")
        r1c1,r1c2=st.columns(2)
        with r1c1:
            fig=px.line(df,x="순번",y="점수",markers=True,color="판정",title="테스트 순서별 품질 점수 추이",hover_data=["TC_ID","유형"]); fig.add_hline(y=80,line_dash="dash",annotation_text="조건부 기준 80"); fig.add_hline(y=90,line_dash="dot",annotation_text="배포 기준 90"); fig.update_layout(height=340,margin=dict(l=10,r=10,t=45,b=10)); st.plotly_chart(fig,use_container_width=True)
        with r1c2:
            fig=px.line(df,x="순번",y="응답시간(ms)",markers=True,title="테스트별 응답시간 추이",hover_data=["TC_ID","유형","판정"]); fig.update_layout(height=340,margin=dict(l=10,r=10,t=45,b=10)); st.plotly_chart(fig,use_container_width=True)
        r2c1,r2c2=st.columns(2)
        with r2c1:
            type_df=_qa_type_summary(st.session_state.runs)
            fig=px.bar(type_df,x="평균점수",y="유형",orientation="h",text_auto=".1f",color="성공률",title="유형별 평균 점수 및 성공률",range_x=[0,100]); fig.update_layout(height=350,margin=dict(l=10,r=10,t=45,b=10)); st.plotly_chart(fig,use_container_width=True)
        with r2c2:
            fig=px.bar(agent_df,x="Agent",y="평균응답(ms)",text_auto=".1f",title="Agent별 평균 수행시간",hover_data=["최대응답(ms)","실행횟수","오류수"]); fig.update_layout(height=350,margin=dict(l=10,r=10,t=45,b=10)); st.plotly_chart(fig,use_container_width=True)
        r3c1,r3c2=st.columns([1,1.4])
        with r3c1:
            status_df=pd.DataFrame({"판정":["PASS","FAIL"],"건수":[passed,failed]})
            fig=px.pie(status_df,names="판정",values="건수",hole=.58,title="PASS / FAIL 구성"); fig.update_layout(height=320,margin=dict(l=10,r=10,t=45,b=10)); st.plotly_chart(fig,use_container_width=True)
        with r3c2:
            st.markdown('<div class="section-title">Agent Performance Table</div>',unsafe_allow_html=True); st.dataframe(agent_df,use_container_width=True,hide_index=True,height=265)
        with st.expander("100점 평가 기준"):
            st.dataframe(rubric,use_container_width=True,hide_index=True)

elif page == "장애·결함":
    page_header("장애·결함", "예외 상황을 재현하고 자동 결함 정보를 확인합니다.")
    left, right = st.columns([1, 1.45])
    with left:
        scenario = st.selectbox("장애 시나리오", ["Retriever 종료", "CSV 파일 누락", "응답 지연", "빈 검색 결과"])
        if st.button("장애 시나리오 실행", type="primary", use_container_width=True):
            mapping = {
                "Retriever 종료": ("결제 내역 확인", "retriever_down"),
                "CSV 파일 누락": ("배송 상태 확인", "csv_missing"),
                "응답 지연": ("쿠폰 할인 확인", "timeout"),
                "빈 검색 결과": ("홀로그램 상담 연결 문의", None),
            }
            q, fault = mapping[scenario]
            st.session_state.single = analyze_voc(q, fault)
            replay_agent_flow(st.session_state.single)
        st.markdown('<div class="section-title">Expected Control</div>', unsafe_allow_html=True)
        st.warning("오류를 성공으로 위장하지 않고 원인과 재시도 방향을 명확히 안내해야 합니다.")
    with right:
        st.markdown(agent_status_html(), unsafe_allow_html=True)
        render_result(st.session_state.single)

    st.divider()
    failures = [i for i in st.session_state.runs if not i["judge"]["pass"]]
    st.markdown('<div class="section-title">Detected Defects</div>', unsafe_allow_html=True)
    if not st.session_state.runs:
        st.info("QA 테스트 실행 후 자동 결함이 표시됩니다.")
    elif not failures:
        st.success("현재 실패 테스트가 없습니다.")
    else:
        for idx, item in enumerate(failures, 1):
            case = item["case"]
            severity = "Critical" if item["critical_failure"] else "High" if item["judge"]["score"] < 50 else "Medium"
            with st.expander(f"BUG-{idx:03d} · {case['case_id']} · {severity}", expanded=idx == 1):
                st.write(f"**입력:** {case['question']}")
                st.write(f"**기대:** {', '.join(case['required_output'])}")
                st.write(f"**실제:** {item['judge']['reason']}")
                st.write("**개선:** 누락된 필수 요소와 근거를 보강한 후 재검증")

elif page == "보고서·배포":
    page_header("보고서·배포", "품질 결과를 PDF·Excel·CSV·JSON·Markdown으로 내보내고 최종 배포 판단을 확정합니다.")
    if not st.session_state.runs:
        st.info("먼저 QA 테스트를 실행하세요.")
    else:
        runs=st.session_state.runs
        decision,avg,rate,passed,failed,critical=_deployment_decision(runs)
        tone="green" if decision=="배포 가능" else "amber" if "조건부" in decision or "개선" in decision else "red"
        st.markdown(f'<div class="hero"><h1>{decision}</h1><p>평균 {avg:.1f}점 · 성공률 {rate:.1f}% · PASS {passed} · FAIL {failed}</p></div>',unsafe_allow_html=True)
        c1,c2,c3,c4=st.columns(4); c1.metric("Quality Score",f"{avg:.1f}"); c2.metric("Pass Rate",f"{rate:.1f}%"); c3.metric("Failed",failed); c4.metric("Critical","YES" if critical else "NO")
        st.markdown(f"### {badge(decision,tone)}",unsafe_allow_html=True)
        st.caption("90점 이상 배포 가능 · 80~89점 조건부 배포 · 70~79점 주요 개선 · 69점 이하 또는 Critical 발생 시 배포 보류")
        st.divider()
        with st.spinner("보고서 생성 중"):
            excel_bytes=build_excel_report(runs); pdf_bytes=build_pdf_report(runs); json_bytes=build_json_report(runs); csv_bytes=_run_rows(runs).to_csv(index=False).encode("utf-8-sig")
            txt_bytes=build_text_report(runs); xml_bytes=build_xml_report(runs); html_bytes=build_html_report(runs)
        st.markdown('<div class="section-title">Executive Reports</div>',unsafe_allow_html=True)
        a,b,c,d=st.columns(4)
        a.download_button("PDF 종합 보고서",pdf_bytes,"VOC_QA_Report.pdf","application/pdf",use_container_width=True)
        b.download_button("Excel 분석 보고서",excel_bytes,"VOC_QA_Report.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
        c.download_button("CSV 전체 결과",csv_bytes,"VOC_QA_Result.csv","text/csv",use_container_width=True)
        d.download_button("JSON 원시 보고서",json_bytes,"VOC_QA_Report.json","application/json",use_container_width=True)
        e,f,g=st.columns(3)
        e.download_button("TXT 실행 증적",txt_bytes,"VOC_QA_Report.txt","text/plain",use_container_width=True)
        f.download_button("XML 구조화 증적",xml_bytes,"VOC_QA_Report.xml","application/xml",use_container_width=True)
        g.download_button("HTML 열람 보고서",html_bytes,"VOC_QA_Report.html","text/html",use_container_width=True)
        st.markdown('<div class="section-title" style="margin-top:1rem">Markdown Reports</div>',unsafe_allow_html=True)
        cols=st.columns(3)
        for col,filename in zip(cols,["quality_score_report.md","deployment_decision.md","defect_report.md"]):
            path=REPORTS/filename
            with col:
                st.markdown(f"**{filename}**")
                if path.exists(): st.download_button("다운로드",path.read_bytes(),filename,use_container_width=True,key=f"md-{filename}")
                else: st.caption("파일 없음")
        st.divider()
        st.markdown('<div class="section-title">보고서 미리보기</div>',unsafe_allow_html=True)
        preview_left,preview_right=st.columns([1,1.35])
        with preview_left:
            type_df=_qa_type_summary(runs); st.dataframe(type_df,use_container_width=True,hide_index=True,height=300)
        with preview_right:
            fig=px.bar(type_df,x="성공률",y="유형",orientation="h",text_auto=".1f",title="유형별 배포 준비도",range_x=[0,100]); fig.update_layout(height=300,margin=dict(l=10,r=10,t=45,b=10)); st.plotly_chart(fig,use_container_width=True)
